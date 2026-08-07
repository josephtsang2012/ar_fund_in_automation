from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Any

import pandas as pd

from .config import AMOUNT_TOLERANCE
from .common import (
    _diff,
    _key_number,
    _normalize_columns,
    _number,
    _plain_header,
    _query_in,
    _read_first_sheet,
    _read_sql_query,
    _resize_plain,
    _sql_values,
    _status,
    _text,
    _trim_table,
    _user_doc,
    _write_plain_sheets,
)

def validate_existing_chain(
    arap_path: str | Path,
    check_line_path: str | Path,
    check_hdr_path: str | Path,
    output_path: str | Path,
    arap_history_line_path: str | Path | None = None,
    header_line_path: str | Path | None = None,
) -> None:
    """
    Validate the confirmed line-factor / effective-header-factor amount chain.

    CHECK_LINE uses an invoice-level factor. CHECK_HDR.CURENCY_FACTOR is an
    effective receipt-level factor calculated from the aggregated line amounts:

        line factor from amount = C_AMOUNT / AMOUNT
        line factor from net    = C_NETAMOUNT / NETAMOUNT

        header ACCOUNT_FACTOR   = ACCOUNT_AMOUNT / AMOUNT
        header CURENCY_FACTOR   = C_AMOUNT / AMOUNT

    ARAP PAID_TOTAL and C_PAID_TOTAL remain cumulative across all historical
    CHECK_LINE rows for the invoice.
    """
    arap = _normalize_columns(_trim_table(_read_first_sheet(arap_path)))
    line = _normalize_columns(_trim_table(_read_first_sheet(check_line_path)))
    hdr = _normalize_columns(_trim_table(_read_first_sheet(check_hdr_path)))
    header_line = (
        line.copy()
        if header_line_path is None
        else _normalize_columns(
            _trim_table(_read_first_sheet(header_line_path))
        )
    )
    arap_history_line = (
        line.copy()
        if arap_history_line_path is None
        else _normalize_columns(
            _trim_table(_read_first_sheet(arap_history_line_path))
        )
    )

    arap["USER_DOC_KEY"] = arap["USER_DOC"].map(_user_doc)
    arap["DOC_CATEGORY_KEY"] = arap["DOC_CATEGORY"].map(_text)
    arap["ACCTNO_KEY"] = arap["ACCTNO"].map(_text)
    arap = arap[arap.get("LINE", 0).fillna(0).eq(0)].copy()

    line["PAY_USER_DOC_KEY"] = line["PAY_USER_DOC"].map(_user_doc)
    line["PAY_DOC_CATEGORY_KEY"] = line["PAY_DOC_CATEGORY"].map(_text)
    line["ACCTNO_KEY"] = line["ACCTNO"].map(_text)
    line["DOC_CATEGORY_KEY"] = line["DOC_CATEGORY"].map(_text)

    header_line["PAY_DOC_CATEGORY_KEY"] = (
        header_line["PAY_DOC_CATEGORY"].map(_text)
    )
    header_line["DOC_CATEGORY_KEY"] = (
        header_line["DOC_CATEGORY"].map(_text)
    )

    arap_history_line["PAY_USER_DOC_KEY"] = (
        arap_history_line["PAY_USER_DOC"].map(_user_doc)
    )
    arap_history_line["PAY_DOC_CATEGORY_KEY"] = (
        arap_history_line["PAY_DOC_CATEGORY"].map(_text)
    )
    arap_history_line["ACCTNO_KEY"] = (
        arap_history_line["ACCTNO"].map(_text)
    )

    hdr["DOC_CATEGORY_KEY"] = hdr["DOC_CATEGORY"].map(_text)

    hdr_index: dict[tuple[Any, str], list[pd.Series]] = {}
    for _, row in hdr.iterrows():
        key = (
            _key_number(row.get("DOC_NO")),
            _text(row.get("DOC_CATEGORY")),
        )
        hdr_index.setdefault(key, []).append(row)

    arap_index: dict[tuple[Any, str, str, str], list[pd.Series]] = {}
    for _, row in arap.iterrows():
        key = (
            _key_number(row.get("DOC_NO")),
            _user_doc(row.get("USER_DOC")),
            _text(row.get("DOC_CATEGORY")),
            _text(row.get("ACCTNO")),
        )
        arap_index.setdefault(key, []).append(row)

    zero_default_columns = [
        "DISCOUNT_DISCREPANCY",
        "NETAMOUNT_DISCREPANCY",
        "CURRENCY_DISCREPANCY",
        "ORIG_DISCOUNT",
        "ORIG_DISCOUNT_DISCREPANCY",
        "ORIG_NETAMOUNT",
        "ORIG_NETAMOUNT_DISCREPANCY",
        "ORIG_AMOUNT",
        "ORIG_CURRENCY_DISCREPANCY",
        "C_DISCOUNT_DISCREPANCY",
        "C_NETAMOUNT_DISCREPANCY",
        "C_CURRENCY_DISCREPANCY",
        "C_ORIG_DISCOUNT",
        "C_ORIG_DISCOUNT_DISCREPANCY",
        "C_ORIG_NETAMOUNT",
        "C_ORIG_NETAMOUNT_DISCREPANCY",
        "C_ORIG_AMOUNT",
        "C_ORIG_CURRENCY_DISCREPANCY",
    ]

    line_rows: list[dict[str, Any]] = []

    for _, line_row in line.iterrows():
        receipt_key = (
            _key_number(line_row.get("DOC_NO")),
            _text(line_row.get("DOC_CATEGORY")),
        )
        invoice_key = (
            _key_number(line_row.get("PAY_DOC_NO")),
            _user_doc(line_row.get("PAY_USER_DOC")),
            _text(line_row.get("PAY_DOC_CATEGORY")),
            _text(line_row.get("ACCTNO")),
        )

        hdr_matches = hdr_index.get(receipt_key, [])
        arap_matches = arap_index.get(invoice_key, [])

        c_amount = _number(line_row.get("C_AMOUNT"))
        c_netamount = _number(line_row.get("C_NETAMOUNT"))
        c_discount = _number(line_row.get("C_DISCOUNT"))
        amount = _number(line_row.get("AMOUNT"))
        netamount = _number(line_row.get("NETAMOUNT"))
        discount = _number(line_row.get("DISCOUNT"))

        expected_c_discount = (
            None
            if c_amount is None or c_netamount is None
            else c_amount - c_netamount
        )
        expected_discount = (
            None
            if amount is None or netamount is None
            else amount - netamount
        )
        factor_from_amount = (
            None
            if c_amount is None or amount in (None, 0)
            else c_amount / amount
        )
        factor_from_netamount = (
            None
            if c_netamount is None or netamount in (None, 0)
            else c_netamount / netamount
        )

        formula_checks = {
            "C_DISCOUNT_EQ_C_AMOUNT_MINUS_C_NETAMOUNT":
                _status(c_discount, expected_c_discount),
            "DISCOUNT_EQ_AMOUNT_MINUS_NETAMOUNT":
                _status(discount, expected_discount),
            "LINE_FACTOR_AMOUNT_VS_NETAMOUNT":
                _status(
                    factor_from_amount,
                    factor_from_netamount,
                    tolerance=1e-6,
                ),
        }

        default_mismatches: list[str] = []
        for column in zero_default_columns:
            if column not in line.columns:
                continue
            value = _number(line_row.get(column))
            if value is not None and abs(value) > AMOUNT_TOLERANCE:
                default_mismatches.append(column)

        errors: list[str] = []
        warnings: list[str] = []
        if len(hdr_matches) == 0:
            errors.append("MISSING_CHECK_HDR")
        elif len(hdr_matches) > 1:
            errors.append("CHECK_HDR_NOT_UNIQUE")

        if len(arap_matches) == 0:
            errors.append("MISSING_ARAP")
        elif len(arap_matches) > 1:
            errors.append("ARAP_NOT_UNIQUE")

        failed_formula_checks = [
            name
            for name, status in formula_checks.items()
            if status == "MISMATCH"
        ]
        if failed_formula_checks:
            errors.append(
                "LINE_AMOUNT_MISMATCH:"
                + ",".join(failed_formula_checks)
            )

        if default_mismatches:
            warnings.append(
                "HISTORICAL_NONZERO_FIELDS:"
                + ",".join(default_mismatches)
            )

        line_rows.append(
            {
                "RECEIPT_DOC_NO": line_row.get("DOC_NO"),
                "LINE": line_row.get("LINE"),
                "PAY_DOC_NO": line_row.get("PAY_DOC_NO"),
                "PAY_USER_DOC": line_row.get("PAY_USER_DOC"),
                "ACCTNO": line_row.get("ACCTNO"),
                "HEADER_MATCH_COUNT": len(hdr_matches),
                "ARAP_MATCH_COUNT": len(arap_matches),
                "ACTUAL_C_AMOUNT": c_amount,
                "ACTUAL_AMOUNT": amount,
                "LINE_FACTOR_FROM_AMOUNT": factor_from_amount,
                "ACTUAL_C_NETAMOUNT": c_netamount,
                "ACTUAL_NETAMOUNT": netamount,
                "LINE_FACTOR_FROM_NETAMOUNT": factor_from_netamount,
                "LINE_FACTOR_DIFF": _diff(
                    factor_from_amount,
                    factor_from_netamount,
                ),
                "EXPECTED_C_DISCOUNT": expected_c_discount,
                "ACTUAL_C_DISCOUNT": c_discount,
                "C_DISCOUNT_DIFF": _diff(
                    c_discount,
                    expected_c_discount,
                ),
                "EXPECTED_DISCOUNT": expected_discount,
                "ACTUAL_DISCOUNT": discount,
                "DISCOUNT_DIFF": _diff(
                    discount,
                    expected_discount,
                ),
                "DEFAULT_ZERO_MISMATCHES":
                    ",".join(default_mismatches),
                "STATUS": "MATCH" if not errors else "MISMATCH",
                "ERRORS": " | ".join(errors),
                "WARNINGS": " | ".join(warnings),
            }
        )

    line_validation = pd.DataFrame(line_rows)

    line_totals = (
        header_line.groupby(
            ["DOC_NO", "DOC_CATEGORY_KEY"],
            dropna=False,
        )
        .agg(
            LINE_COUNT=("LINE", "count"),
            EXPECTED_APPLIED=("AMOUNT", "sum"),
            EXPECTED_AMOUNT=("NETAMOUNT", "sum"),
            EXPECTED_DISCOUNT=("DISCOUNT", "sum"),
            EXPECTED_C_APPLIED=("C_AMOUNT", "sum"),
            EXPECTED_C_AMOUNT=("C_NETAMOUNT", "sum"),
            EXPECTED_C_DISCOUNT=("C_DISCOUNT", "sum"),
        )
        .reset_index()
    )

    receipt_categories = (
        header_line.groupby(
            ["DOC_NO", "DOC_CATEGORY_KEY"],
            dropna=False,
        )["PAY_DOC_CATEGORY_KEY"]
        .agg(
            lambda values: ",".join(
                sorted(
                    {
                        _text(value) or "<BLANK>"
                        for value in values
                    }
                )
            )
        )
        .reset_index(name="PAY_DOC_CATEGORIES")
    )
    line_totals = line_totals.merge(
        receipt_categories,
        how="left",
        on=["DOC_NO", "DOC_CATEGORY_KEY"],
    )

    header = hdr.merge(
        line_totals,
        how="left",
        on=["DOC_NO", "DOC_CATEGORY_KEY"],
    )

    header_rows: list[dict[str, Any]] = []

    for _, row in header.iterrows():
        account_amount = _number(row.get("ACCOUNT_AMOUNT"))
        amount = _number(row.get("AMOUNT"))
        c_amount = _number(row.get("C_AMOUNT"))

        expected_account_factor = (
            None
            if account_amount is None or amount in (None, 0)
            else account_amount / amount
        )
        expected_currency_factor = (
            None
            if c_amount is None or amount in (None, 0)
            else c_amount / amount
        )

        checks = {
            "APPLIED_EQ_SUM_LINE_AMOUNT":
                _status(row.get("APPLIED"), row.get("EXPECTED_APPLIED")),
            "AMOUNT_EQ_SUM_LINE_NETAMOUNT":
                _status(row.get("AMOUNT"), row.get("EXPECTED_AMOUNT")),
            "NETAMOUNT_EQ_AMOUNT":
                _status(row.get("NETAMOUNT"), row.get("AMOUNT")),
            "DISCOUNT_EQ_SUM_LINE_DISCOUNT":
                _status(row.get("DISCOUNT"), row.get("EXPECTED_DISCOUNT")),
            "C_APPLIED_EQ_SUM_LINE_C_AMOUNT":
                _status(row.get("C_APPLIED"), row.get("EXPECTED_C_APPLIED")),
            "C_AMOUNT_EQ_SUM_LINE_C_NETAMOUNT":
                _status(row.get("C_AMOUNT"), row.get("EXPECTED_C_AMOUNT")),
            "C_NETAMOUNT_EQ_C_AMOUNT":
                _status(row.get("C_NETAMOUNT"), row.get("C_AMOUNT")),
            "C_DISCOUNT_EQ_SUM_LINE_C_DISCOUNT":
                _status(
                    row.get("C_DISCOUNT"),
                    row.get("EXPECTED_C_DISCOUNT"),
                ),
            "ACCOUNT_FACTOR_EQ_ACCOUNT_AMOUNT_DIV_AMOUNT":
                _status(
                    row.get("ACCOUNT_FACTOR"),
                    expected_account_factor,
                    tolerance=1e-6,
                ),
            "CURENCY_FACTOR_EQ_C_AMOUNT_DIV_AMOUNT":
                _status(
                    row.get("CURENCY_FACTOR"),
                    expected_currency_factor,
                    tolerance=1e-6,
                ),
        }

        failed_checks = [
            name
            for name, status in checks.items()
            if status == "MISMATCH"
        ]

        category_values = {
            value
            for value in _text(
                row.get("PAY_DOC_CATEGORIES")
            ).split(",")
            if value
        }
        non_iv_categories = sorted(
            value for value in category_values if value != "IV"
        )

        header_factor = _number(row.get("CURENCY_FACTOR"))
        applied_factor_residual = (
            None
            if header_factor in (None, 0)
            else _diff(
                row.get("APPLIED"),
                _number(row.get("C_APPLIED"), 0)
                / header_factor,
            )
        )
        amount_factor_residual = (
            None
            if header_factor in (None, 0)
            else _diff(
                row.get("AMOUNT"),
                _number(row.get("C_AMOUNT"), 0)
                / header_factor,
            )
        )

        factor_precision_checks = {
            "APPLIED_EQ_SUM_LINE_AMOUNT",
            "AMOUNT_EQ_SUM_LINE_NETAMOUNT",
        }
        factor_precision_only = (
            bool(failed_checks)
            and set(failed_checks).issubset(factor_precision_checks)
            and applied_factor_residual is not None
            and amount_factor_residual is not None
            and abs(applied_factor_residual) <= 0.000001
            and abs(amount_factor_residual) <= 0.000001
        )

        errors: list[str] = []
        warnings: list[str] = []
        if pd.isna(row.get("LINE_COUNT")):
            errors.append("NO_CHECK_LINE")

        if non_iv_categories:
            status = "OUT_OF_SCOPE"
            classification = (
                "OUT_OF_SCOPE_NON_IV_LINES:"
                + ",".join(non_iv_categories)
            )
            warnings.append(classification)
            if failed_checks:
                warnings.append(
                    "HISTORICAL_HEADER_DIFFERENCE:"
                    + ",".join(failed_checks)
                )
        elif factor_precision_only:
            status = "MATCH_WITHIN_FACTOR_PRECISION"
            classification = "FACTOR_PRECISION_ONLY"
            warnings.append(
                "HEADER_BASE_AMOUNT_FACTOR_PRECISION:"
                + ",".join(failed_checks)
            )
        elif failed_checks:
            status = "MISMATCH"
            classification = "UNEXPLAINED_CORE_MISMATCH"
            errors.append(
                "HDR_AMOUNT_MISMATCH:"
                + ",".join(failed_checks)
            )
        else:
            status = "MATCH"
            classification = "STANDARD_IV_MATCH"

        header_rows.append(
            {
                "DOC_NO": row.get("DOC_NO"),
                "CHECK_NO": row.get("CHECK_NO"),
                "LINE_COUNT": row.get("LINE_COUNT"),
                "EXPECTED_APPLIED": row.get("EXPECTED_APPLIED"),
                "ACTUAL_APPLIED": row.get("APPLIED"),
                "APPLIED_DIFF": _diff(
                    row.get("APPLIED"),
                    row.get("EXPECTED_APPLIED"),
                ),
                "EXPECTED_AMOUNT": row.get("EXPECTED_AMOUNT"),
                "ACTUAL_AMOUNT": row.get("AMOUNT"),
                "AMOUNT_DIFF": _diff(
                    row.get("AMOUNT"),
                    row.get("EXPECTED_AMOUNT"),
                ),
                "ACTUAL_NETAMOUNT": row.get("NETAMOUNT"),
                "NETAMOUNT_MINUS_AMOUNT": _diff(
                    row.get("NETAMOUNT"),
                    row.get("AMOUNT"),
                ),
                "EXPECTED_DISCOUNT": row.get("EXPECTED_DISCOUNT"),
                "ACTUAL_DISCOUNT": row.get("DISCOUNT"),
                "DISCOUNT_DIFF": _diff(
                    row.get("DISCOUNT"),
                    row.get("EXPECTED_DISCOUNT"),
                ),
                "EXPECTED_C_APPLIED": row.get("EXPECTED_C_APPLIED"),
                "ACTUAL_C_APPLIED": row.get("C_APPLIED"),
                "C_APPLIED_DIFF": _diff(
                    row.get("C_APPLIED"),
                    row.get("EXPECTED_C_APPLIED"),
                ),
                "EXPECTED_C_AMOUNT": row.get("EXPECTED_C_AMOUNT"),
                "ACTUAL_C_AMOUNT": row.get("C_AMOUNT"),
                "C_AMOUNT_DIFF": _diff(
                    row.get("C_AMOUNT"),
                    row.get("EXPECTED_C_AMOUNT"),
                ),
                "ACTUAL_C_NETAMOUNT": row.get("C_NETAMOUNT"),
                "C_NETAMOUNT_MINUS_C_AMOUNT": _diff(
                    row.get("C_NETAMOUNT"),
                    row.get("C_AMOUNT"),
                ),
                "EXPECTED_C_DISCOUNT": row.get("EXPECTED_C_DISCOUNT"),
                "ACTUAL_C_DISCOUNT": row.get("C_DISCOUNT"),
                "C_DISCOUNT_DIFF": _diff(
                    row.get("C_DISCOUNT"),
                    row.get("EXPECTED_C_DISCOUNT"),
                ),
                "EXPECTED_ACCOUNT_FACTOR": expected_account_factor,
                "ACTUAL_ACCOUNT_FACTOR": row.get("ACCOUNT_FACTOR"),
                "ACCOUNT_FACTOR_DIFF": _diff(
                    row.get("ACCOUNT_FACTOR"),
                    expected_account_factor,
                ),
                "EXPECTED_CURENCY_FACTOR": expected_currency_factor,
                "ACTUAL_CURENCY_FACTOR": row.get("CURENCY_FACTOR"),
                "CURENCY_FACTOR_DIFF": _diff(
                    row.get("CURENCY_FACTOR"),
                    expected_currency_factor,
                ),
                "PAY_DOC_CATEGORIES": row.get("PAY_DOC_CATEGORIES"),
                "NON_IV_CATEGORIES": ",".join(non_iv_categories),
                "APPLIED_FACTOR_RESIDUAL": applied_factor_residual,
                "AMOUNT_FACTOR_RESIDUAL": amount_factor_residual,
                "CLASSIFICATION": classification,
                "STATUS": status,
                "ERRORS": " | ".join(errors),
                "WARNINGS": " | ".join(warnings),
            }
        )

    header_validation = pd.DataFrame(header_rows)

    invoice_line_totals = (
        arap_history_line.groupby(
            [
                "PAY_DOC_NO",
                "PAY_USER_DOC_KEY",
                "PAY_DOC_CATEGORY_KEY",
                "ACCTNO_KEY",
            ],
            dropna=False,
        )
        .agg(
            PAYMENT_LINE_COUNT=("LINE", "count"),
            EXPECTED_PAID_TOTAL=("NETAMOUNT", "sum"),
            EXPECTED_C_PAID_TOTAL=("C_NETAMOUNT", "sum"),
        )
        .reset_index()
    )

    invoice_totals_index: dict[
        tuple[Any, str, str, str],
        pd.Series,
    ] = {}
    for _, row in invoice_line_totals.iterrows():
        key = (
            _key_number(row.get("PAY_DOC_NO")),
            _user_doc(row.get("PAY_USER_DOC_KEY")),
            _text(row.get("PAY_DOC_CATEGORY_KEY")),
            _text(row.get("ACCTNO_KEY")),
        )
        invoice_totals_index[key] = row

    arap_rows: list[dict[str, Any]] = []

    for _, arap_row in arap.iterrows():
        key = (
            _key_number(arap_row.get("DOC_NO")),
            _user_doc(arap_row.get("USER_DOC")),
            _text(arap_row.get("DOC_CATEGORY")),
            _text(arap_row.get("ACCTNO")),
        )
        totals = invoice_totals_index.get(key)

        if totals is None:
            arap_rows.append(
                {
                    "DOC_NO": arap_row.get("DOC_NO"),
                    "USER_DOC": arap_row.get("USER_DOC"),
                    "DOC_CATEGORY": arap_row.get("DOC_CATEGORY"),
                    "ACCTNO": arap_row.get("ACCTNO"),
                    "PAYMENT_LINE_COUNT": 0,
                    "EXPECTED_PAID_TOTAL": None,
                    "ACTUAL_PAID_TOTAL": arap_row.get("PAID_TOTAL"),
                    "PAID_TOTAL_DIFF": None,
                    "EXPECTED_C_PAID_TOTAL": None,
                    "ACTUAL_C_PAID_TOTAL": arap_row.get("C_PAID_TOTAL"),
                    "C_PAID_TOTAL_DIFF": None,
                    "STATUS": "NOT_TESTED",
                    "ERRORS": "",
                    "WARNINGS":
                        "NO_MATCHING_CHECK_LINE_IN_HISTORY_SCOPE",
                    "COVERAGE_NOTE": (
                        "Diagnostic only. ARAP totals may include reversals, "
                        "voids, or payment document categories outside the "
                        "confirmed AR/R2 generation scope."
                    ),
                }
            )
            continue

        expected_paid_total = totals.get("EXPECTED_PAID_TOTAL")
        expected_c_paid_total = totals.get("EXPECTED_C_PAID_TOTAL")

        actual_paid_total = _number(arap_row.get("PAID_TOTAL"))
        actual_c_paid_total = _number(
            arap_row.get("C_PAID_TOTAL")
        )
        paid_diff = (
            None
            if actual_paid_total is None
            or _number(expected_paid_total) is None
            else actual_paid_total - float(expected_paid_total)
        )
        c_paid_diff = (
            None
            if actual_c_paid_total is None
            or _number(expected_c_paid_total) is None
            else actual_c_paid_total - float(expected_c_paid_total)
        )
        arap_factor = abs(
            _number(arap_row.get("CURENCY_FACTOR"), 1.0) or 1.0
        )
        base_tolerance = AMOUNT_TOLERANCE
        converted_tolerance = max(
            AMOUNT_TOLERANCE,
            arap_factor * AMOUNT_TOLERANCE,
        )

        standard_base_match = (
            paid_diff is not None
            and abs(paid_diff) <= AMOUNT_TOLERANCE
        )
        standard_converted_match = (
            c_paid_diff is not None
            and abs(c_paid_diff) <= AMOUNT_TOLERANCE
        )
        scaled_rounding_match = (
            paid_diff is not None
            and c_paid_diff is not None
            and abs(paid_diff) <= base_tolerance
            and abs(c_paid_diff) <= converted_tolerance
        )

        if standard_base_match and standard_converted_match:
            status = "MATCH"
            classification = "EXACT_WITHIN_STANDARD_TOLERANCE"
            warnings: list[str] = []
        elif scaled_rounding_match:
            status = "MATCH_WITHIN_ROUNDING"
            classification = "FACTOR_SCALED_ROUNDING"
            warnings = [
                "ARAP_CUMULATIVE_ROUNDING_ONLY:"
                f"BASE_DIFF={round(paid_diff, 12)},"
                f"C_DIFF={round(c_paid_diff, 12)},"
                f"FACTOR={arap_factor}"
            ]
        else:
            failed_checks: list[str] = []
            if not standard_base_match:
                failed_checks.append(
                    "PAID_TOTAL_EQ_SUM_HISTORY_LINE_NETAMOUNT"
                )
            if not standard_converted_match:
                failed_checks.append(
                    "C_PAID_TOTAL_EQ_SUM_HISTORY_LINE_C_NETAMOUNT"
                )
            status = "REVIEW"
            classification = "UNCONFIRMED_CUMULATIVE_RULE"
            warnings = [
                "ARAP_CUMULATIVE_REVIEW:"
                + ",".join(failed_checks)
            ]

        arap_rows.append(
            {
                "DOC_NO": arap_row.get("DOC_NO"),
                "USER_DOC": arap_row.get("USER_DOC"),
                "DOC_CATEGORY": arap_row.get("DOC_CATEGORY"),
                "ACCTNO": arap_row.get("ACCTNO"),
                "PAYMENT_LINE_COUNT": totals.get("PAYMENT_LINE_COUNT"),
                "EXPECTED_PAID_TOTAL": expected_paid_total,
                "ACTUAL_PAID_TOTAL": arap_row.get("PAID_TOTAL"),
                "PAID_TOTAL_DIFF": (
                    None if paid_diff is None else round(paid_diff, 10)
                ),
                "EXPECTED_C_PAID_TOTAL": expected_c_paid_total,
                "ACTUAL_C_PAID_TOTAL": arap_row.get("C_PAID_TOTAL"),
                "C_PAID_TOTAL_DIFF": (
                    None if c_paid_diff is None else round(c_paid_diff, 10)
                ),
                "ARAP_CURENCY_FACTOR": arap_factor,
                "BASE_TOLERANCE": base_tolerance,
                "CONVERTED_TOLERANCE": converted_tolerance,
                "CLASSIFICATION": classification,
                "STATUS": status,
                "ERRORS": "",
                "WARNINGS": " | ".join(warnings),
                "COVERAGE_NOTE": (
                    "Diagnostic only. Exact equality is not used to gate "
                    "the confirmed CHECK_HDR/CHECK_LINE generation logic."
                ),
            }
        )

    arap_validation = pd.DataFrame(arap_rows)

    def _status_count(
        frame: pd.DataFrame,
        status: str,
    ) -> int:
        if frame.empty or "STATUS" not in frame.columns:
            return 0
        return int(frame["STATUS"].eq(status).sum())

    validation_summary = pd.DataFrame(
        [
            {
                "AREA": "CHECK_LINE_CORE",
                "GATING": "YES",
                "TOTAL": len(line_validation),
                "MATCH": _status_count(line_validation, "MATCH"),
                "MISMATCH": _status_count(
                    line_validation,
                    "MISMATCH",
                ),
                "REVIEW": 0,
                "NOT_TESTED": 0,
                "INTERPRETATION": (
                    "Formula and relationship validation. Historical "
                    "nonzero discrepancy/original fields are warnings."
                ),
            },
            {
                "AREA": "CHECK_HDR_CORE",
                "GATING": "YES",
                "TOTAL": len(header_validation),
                "MATCH": _status_count(header_validation, "MATCH"),
                "MISMATCH": _status_count(
                    header_validation,
                    "MISMATCH",
                ),
                "REVIEW": 0,
                "MATCH_WITHIN_FACTOR_PRECISION": _status_count(
                    header_validation,
                    "MATCH_WITHIN_FACTOR_PRECISION",
                ),
                "OUT_OF_SCOPE": _status_count(
                    header_validation,
                    "OUT_OF_SCOPE",
                ),
                "NOT_TESTED": 0,
                "INTERPRETATION": (
                    "Receipt totals use all CHECK_LINE categories. "
                    "AD/MS receipts are classified out of scope; "
                    "small base-side factor residuals are classified "
                    "as precision matches."
                ),
            },
            {
                "AREA": "ARAP_CUMULATIVE_DIAGNOSTIC",
                "GATING": "NO",
                "TOTAL": len(arap_validation),
                "MATCH": _status_count(arap_validation, "MATCH"),
                "MISMATCH": 0,
                "REVIEW": _status_count(arap_validation, "REVIEW"),
                "MATCH_WITHIN_ROUNDING": _status_count(
                    arap_validation,
                    "MATCH_WITHIN_ROUNDING",
                ),
                "NOT_TESTED": _status_count(
                    arap_validation,
                    "NOT_TESTED",
                ),
                "INTERPRETATION": (
                    "Diagnostic only until the full inclusion rules for "
                    "voids, reversals, and other payment categories are "
                    "confirmed."
                ),
            },
        ]
    )

    _write_plain_sheets(
        output_path,
        {
            "Validation_Summary": validation_summary,
            "Header_Validation": header_validation,
            "Line_Validation": line_validation,
            "ARAP_Diagnostic": arap_validation,
        },
    )

def _parse_doc_no_list(value: str | None) -> list[int]:
    if not value:
        return []

    result: list[int] = []
    for token in value.split(","):
        token = token.strip()
        if not token:
            continue
        try:
            result.append(int(token))
        except ValueError as exc:
            raise ValueError(
                f"Invalid DOC_NO in --doc-no: {token}"
            ) from exc
    return result


def _historical_scope_doc_nos(
    connection: Any,
    top: int,
    doc_nos: list[int],
    date_from: str | None,
    date_to: str | None,
    currency_mode: str,
) -> list[int]:
    if doc_nos:
        return sorted(set(doc_nos))

    if top <= 0:
        raise ValueError("--top must be greater than 0.")

    where = [
        "RTRIM(H.DOC_TYPE) = 'AR'",
        "RTRIM(H.DOC_CATEGORY) = 'R2'",
        "EXISTS ("
        " SELECT 1"
        " FROM dbo.CHECK_LINE L"
        " WHERE L.DOC_NO = H.DOC_NO"
        "   AND RTRIM(L.DOC_CATEGORY) = RTRIM(H.DOC_CATEGORY)"
        "   AND RTRIM(L.PAY_DOC_CATEGORY) = 'IV'"
        ")",
    ]
    params: list[Any] = []

    if date_from:
        where.append("H.CHECK_DATE >= ?")
        params.append(date_from)
    if date_to:
        where.append("H.CHECK_DATE < DATEADD(day, 1, CAST(? AS date))")
        params.append(date_to)

    normalized_account = (
        "CASE "
        "WHEN UPPER(RTRIM(H.ACCOUNT_CURRENCY)) IN ('RMB','CNY') "
        "THEN 'CNY' "
        "ELSE UPPER(RTRIM(H.ACCOUNT_CURRENCY)) END"
    )
    normalized_conv = (
        "CASE "
        "WHEN UPPER(RTRIM(H.CURENCY_CONV)) IN ('RMB','CNY') "
        "THEN 'CNY' "
        "ELSE UPPER(RTRIM(H.CURENCY_CONV)) END"
    )

    if currency_mode == "same":
        where.append(f"{normalized_account} = {normalized_conv}")
    elif currency_mode == "cross":
        where.append(f"{normalized_account} <> {normalized_conv}")

    sql = f"""
        SELECT TOP {int(top)}
            H.DOC_NO
        FROM dbo.CHECK_HDR H
        WHERE {" AND ".join(where)}
        ORDER BY H.CHECK_DATE DESC, H.DOC_NO DESC
    """
    scope = _read_sql_query(connection, sql, params)
    if scope.empty:
        return []

    return [
        int(value)
        for value in scope["DOC_NO"].dropna().tolist()
    ]


def validate_existing_chain_from_db(
    output_path: str | Path,
    connection_string: str,
    top: int = 1000,
    doc_no: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    currency_mode: str = "all",
) -> None:
    """
    Validate historical records directly from SQL Server.

    Core CHECK_HDR and CHECK_LINE validation uses the complete line set for
    each selected receipt. ARAP cumulative comparison uses a separate,
    broader payment-history query and is reported as diagnostic only.
    """
    try:
        import pyodbc
    except ImportError as exc:
        raise RuntimeError(
            "pyodbc is required. Install it with: pip install pyodbc"
        ) from exc

    requested_doc_nos = _parse_doc_no_list(doc_no)

    with pyodbc.connect(connection_string) as connection:
        focus_doc_nos = _historical_scope_doc_nos(
            connection=connection,
            top=top,
            doc_nos=requested_doc_nos,
            date_from=date_from,
            date_to=date_to,
            currency_mode=currency_mode,
        )

        if not focus_doc_nos:
            raise ValueError(
                "No supported AR/R2/IV historical receipts matched "
                "the requested DB scope."
            )

        # Retrieve every R2 line for Header aggregation/classification.
        # The confirmed generation scope remains PAY_DOC_CATEGORY='IV',
        # while AD/MS lines are reported as out-of-scope historical cases.
        focus_all_lines = _query_in(
            connection,
            "dbo.CHECK_LINE",
            "DOC_NO",
            focus_doc_nos,
            extra_where="RTRIM(DOC_CATEGORY) = 'R2'",
        )
        focus_all_lines = _normalize_columns(focus_all_lines)
        focus_lines = focus_all_lines[
            focus_all_lines["PAY_DOC_CATEGORY"]
            .map(_text)
            .eq("IV")
        ].copy()

        focus_hdr = _query_in(
            connection,
            "dbo.CHECK_HDR",
            "DOC_NO",
            focus_doc_nos,
            extra_where=(
                "RTRIM(DOC_TYPE) = 'AR' "
                "AND RTRIM(DOC_CATEGORY) = 'R2'"
            ),
        )
        focus_hdr = _normalize_columns(focus_hdr)

        pay_doc_nos = _sql_values(
            focus_lines.get(
                "PAY_DOC_NO",
                pd.Series(dtype=object),
            )
        )
        if not pay_doc_nos:
            raise ValueError(
                "The selected receipts contain no IV CHECK_LINE rows."
            )

        arap = _query_in(
            connection,
            "dbo.ACCOUNT_AR_AP",
            "DOC_NO",
            pay_doc_nos,
            extra_where=(
                "ISNULL(LINE, 0) = 0 "
                "AND RTRIM(DOC_CATEGORY) = 'IV'"
            ),
        )
        arap = _normalize_columns(arap)

        # Broad invoice payment history for ARAP diagnostics. Do not restrict
        # CHECK_LINE.DOC_CATEGORY to R2 here because ARAP totals may reflect
        # other payment/reversal categories.
        history_lines = _query_in(
            connection,
            "dbo.CHECK_LINE",
            "PAY_DOC_NO",
            pay_doc_nos,
            extra_where="RTRIM(PAY_DOC_CATEGORY) = 'IV'",
        )
        history_lines = _normalize_columns(history_lines)

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_dir_path = Path(temp_dir)
        arap_path = temp_dir_path / "arap.xlsx"
        focus_line_path = temp_dir_path / "focus_check_line.xlsx"
        all_line_path = temp_dir_path / "all_receipt_lines.xlsx"
        focus_hdr_path = temp_dir_path / "focus_check_hdr.xlsx"
        history_line_path = temp_dir_path / "arap_history_line.xlsx"

        arap.to_excel(arap_path, index=False)
        focus_lines.to_excel(focus_line_path, index=False)
        focus_all_lines.to_excel(all_line_path, index=False)
        focus_hdr.to_excel(focus_hdr_path, index=False)
        history_lines.to_excel(history_line_path, index=False)

        validate_existing_chain(
            arap_path=arap_path,
            check_line_path=focus_line_path,
            check_hdr_path=focus_hdr_path,
            output_path=output_path,
            arap_history_line_path=history_line_path,
            header_line_path=all_line_path,
        )

    scope_rows = pd.DataFrame(
        {
            "FOCUS_DOC_NO": focus_doc_nos,
            "CURRENCY_MODE": currency_mode,
            "DATE_FROM": date_from,
            "DATE_TO": date_to,
            "CORE_LINE_SCOPE": "COMPLETE_SELECTED_RECEIPT",
            "ARAP_SCOPE": "ALL_IV_PAYMENT_LINES_FOUND_FOR_FOCUS_INVOICES",
        }
    )

    from openpyxl import load_workbook

    workbook = load_workbook(output_path)
    if "DB_Scope" in workbook.sheetnames:
        del workbook["DB_Scope"]
    worksheet = workbook.create_sheet("DB_Scope")

    for column_index, column_name in enumerate(
        scope_rows.columns,
        start=1,
    ):
        worksheet.cell(1, column_index, column_name)

    for row_index, row in enumerate(
        scope_rows.itertuples(index=False),
        start=2,
    ):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row_index, column_index, value)

    worksheet.freeze_panes = "A2"
    _plain_header(worksheet, 1)
    _resize_plain(worksheet)
    workbook.save(output_path)
