from __future__ import annotations

import argparse
import os
import tempfile
from dataclasses import dataclass
from datetime import timedelta
from pathlib import Path
from copy import copy
from typing import Any, Iterable

import pandas as pd


# -----------------------------------------------------------------------------
# Environment / SQL Server connection helpers
# -----------------------------------------------------------------------------

def _load_env_file(path: str | Path) -> None:
    """Load a simple KEY=VALUE .env file without adding another dependency."""
    env_path = Path(path)
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()

        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
            value = value[1:-1]

        if key:
            # Existing operating-system variables take priority over the file.
            os.environ.setdefault(key, value)


def _odbc_value(value: str) -> str:
    """Safely quote a SQL Server ODBC connection-string value."""
    return "{" + str(value).replace("}", "}}") + "}"


def _connection_string_from_env(env_file: str | Path = ".env") -> str:
    """
    Load the SQL Server connection string from .env.

    Preferred:
        DATABASE_CONNECTION_STRING=DRIVER={SQL Server};SERVER=...;DATABASE=...;UID=...;PWD=...

    Fallback:
        Build it from separate P2000_DB_* variables.
    """
    _load_env_file(env_file)

    # Prefer the exact full connection string already known to work locally.
    full_connection_string = (
        os.getenv("DATABASE_CONNECTION_STRING", "").strip()
        or os.getenv("P2000_DATABASE_CONNECTION_STRING", "").strip()
    )
    if full_connection_string:
        return full_connection_string.rstrip(";") + ";"

    driver = os.getenv("P2000_DB_DRIVER", "SQL Server").strip()
    server = os.getenv("P2000_DB_SERVER", "").strip()
    database = os.getenv("P2000_DB_DATABASE", "").strip()
    username = os.getenv("P2000_DB_USER", "").strip()
    password = os.getenv("P2000_DB_PASSWORD", "")
    timeout = os.getenv("P2000_DB_TIMEOUT", "30").strip()

    missing = [
        name
        for name, value in {
            "P2000_DB_SERVER": server,
            "P2000_DB_DATABASE": database,
            "P2000_DB_USER": username,
            "P2000_DB_PASSWORD": password,
        }.items()
        if not value
    ]
    if missing:
        raise ValueError(
            "Missing database settings: " + ", ".join(missing)
            + f". Add DATABASE_CONNECTION_STRING or the separate P2000_DB_* values to {env_file}."
        )

    return ";".join(
        [
            f"DRIVER={_odbc_value(driver)}",
            f"SERVER={_odbc_value(server)}",
            f"DATABASE={_odbc_value(database)}",
            f"UID={_odbc_value(username)}",
            f"PWD={_odbc_value(password)}",
            f"Connection Timeout={timeout}",
        ]
    ) + ";"


AMOUNT_TOLERANCE = 0.05
FACTOR_TOLERANCE = 1e-8

COUNTER_ROLE_MAP = {
    "CHECK_DOC_NO": ("CHECK", "D"),
    "R2_CHECK_NO": ("R2", "D"),
    "ACRN_CHECK_COUNTER": ("ACRN", "D"),
    "BC_BATCH_NO": ("BC", "D"),
    "R2_LINE_NO": ("R2", "L"),
}

HDR_SCHEMA = ['TableName', 'DOC_NO', 'DOC_TYPE', 'DOC_CATEGORY', 'CHECK_NO', 'DOC_STATUS', 'CHECK_COUNTER', 'CRDT_DOC_NO', 'CRDT_LINE', 'CRDT_DOC_TYPE', 'CRDT_DOC_CATEGORY', 'CRDT_CHECK_NO', 'MANUAL_AUTO', 'AUTO_BATCH', 'STATUS', 'BANK_ID', 'ACCOUNT_NO', 'CHECKBOOK_NO', 'ACCTNO', 'SUBC', 'PAYTO', 'PAYTO_SUBC', 'PAYTO_CCODE', 'CUST_VEND', 'CHECK_DATE', 'PAYMENT_DATE', 'NOTE', 'PAYEE', 'GL_CODE', 'COMPANY', 'DIVISION', 'DEPART', 'LAST_STATEMENT_BALANCE', 'INFO_BANK_ID', 'INFO_BANK_NAME', 'INFO_SWIFT', 'INFO_ACCOUNT_NO', 'INFO_BRANCH_NO', 'WIRE_PROCESS_DATE', 'WIRE_FILE_NAME', 'CC_CREDITCARD', 'CC_NAME', 'CC_NUMBER', 'CC_EXP', 'PRINT_DATE', 'POST_GL_DATE', 'POST_GL', 'BATCH_NO', 'BATCH_DATE', 'STATEMENT_DOC_NO', 'DEPOSIT_DOC_NO', 'CLEARED', 'CLEARED_BY', 'CLEARED_ON', 'CLEARED_CURRENCY_STATE', 'PRINTED_ON', 'CURENCY_STATE', 'CURENCY_BASE', 'CURENCY_CONV', 'CURENCY_FACTOR', 'CURENCY_DATE', 'DISCOUNT', 'DISCOUNT_DISCREPANCY', 'NETAMOUNT', 'NETAMOUNT_DISCREPANCY', 'AMOUNT', 'APPLIED', 'CURRENCY_DISCREPANCY', 'OPEN_BALANCE', 'C_AMOUNT', 'C_DISCOUNT', 'C_DISCOUNT_DISCREPANCY', 'C_NETAMOUNT', 'C_NETAMOUNT_DISCREPANCY', 'C_APPLIED', 'C_CURRENCY_DISCREPANCY', 'C_OPEN_BALANCE', 'ORIG_DISCOUNT', 'ORIG_DISCOUNT_DISCREPANCY', 'ORIG_NETAMOUNT', 'ORIG_NETAMOUNT_DISCREPANCY', 'ORIG_AMOUNT', 'ORIG_APPLIED', 'ORIG_OPEN_BALANCE', 'C_ORIG_AMOUNT', 'C_ORIG_DISCOUNT', 'C_ORIG_DISCOUNT_DISCREPANCY', 'C_ORIG_NETAMOUNT', 'C_ORIG_NETAMOUNT_DISCREPANCY', 'C_ORIG_APPLIED', 'C_ORIG_OPEN_BALANCE', 'PAID_BY_ACH', 'ACH_ACCOUNT_NO', 'ACH_ACCOUNT_NAME', 'ACH_TRANSACTION_CODE', 'ACH_BANK_NAME', 'ACH_ROUTING_NO', 'BRANCH_NO', 'BANK_NO', 'BANK_NAME', 'ACCOUNT_CURRENCY', 'ACCOUNT_AMOUNT', 'ACCOUNT_FACTOR', 'ACCOUNT_FACTOR_DATE', 'TRANSFER_GL_CODE', 'ACCOUNT_ORIG_AMOUNT', 'DOCUMENT_REFERENCE', 'MEMO', 'SIGN_1', 'SIGN_2', 'SIGN_3', 'SIGN_4', 'SIGN_5', 'SIGN_1_APR', 'SIGN_2_APR', 'SIGN_3_APR', 'SIGN_4_APR', 'SIGN_5_APR', 'SIGN_1_DATE', 'SIGN_2_DATE', 'SIGN_3_DATE', 'SIGN_4_DATE', 'SIGN_5_DATE', 'APPLIED_BY', 'ACH_PAYMENT_TYPE', 'RECLOCK', 'ADDED_USR', 'ADDED_DTE', 'UPDATED_USR', 'UPDATED_DTE']
LINE_SCHEMA = ['TableName', 'DOC_NO', 'DOC_TYPE', 'DOC_CATEGORY', 'DOC_STATUS', 'CHECK_NO', 'PAY_USER_DOC', 'PAY_DOC_CATEGORY', 'LINE', 'PAY_DOC_NO', 'PAY_APPLY', 'PAY_DOC_DATE', 'PAY_TYPE', 'DISCOUNT', 'DISCOUNT_DISCREPANCY', 'NETAMOUNT', 'NETAMOUNT_DISCREPANCY', 'AMOUNT', 'CURRENCY_DISCREPANCY', 'ORIG_DISCOUNT', 'ORIG_DISCOUNT_DISCREPANCY', 'ORIG_NETAMOUNT', 'ORIG_NETAMOUNT_DISCREPANCY', 'ORIG_AMOUNT', 'ORIG_CURRENCY_DISCREPANCY', 'ACCTNO', 'SUBC', 'GL_CODE_AMT', 'GL_CODE_DISC', 'GL_CODE_AMT_DISCREPANCY', 'GL_CODE_DISC_DISCREPANCY_DB', 'GL_CODE_DISC_DISCREPANCY_CR', 'COMPANYNO', 'DIVISION', 'DEPART', 'C_DISCOUNT', 'C_DISCOUNT_DISCREPANCY', 'C_NETAMOUNT', 'C_NETAMOUNT_DISCREPANCY', 'C_AMOUNT', 'C_CURRENCY_DISCREPANCY', 'C_ORIG_DISCOUNT', 'C_ORIG_DISCOUNT_DISCREPANCY', 'C_ORIG_NETAMOUNT', 'C_ORIG_NETAMOUNT_DISCREPANCY', 'C_ORIG_AMOUNT', 'C_ORIG_CURRENCY_DISCREPANCY', 'FORM_1099TYPE', 'TRANSACTION_ID', 'TRANSACTION_ID_SOURCE', 'STATEMENT_DOC_NO', 'CLEARED', 'CLEARED_BY', 'CLEARED_ON', 'RECLOCK', 'ADDED_USR', 'ADDED_DTE', 'UPDATED_USR', 'UPDATED_DTE']
ARAP_SCHEMA = ['TableName', 'DOC_NO', 'LINE', 'DOC_CATEGORY', 'USER_DOC', 'DOC_TYPE', 'AR_AP', 'DOC_STATUS', 'PARENT_DOC_NO', 'PARENT_LINE', 'PARENT_DOCCATEGORY', 'ACCTNO', 'SUBC', 'CUST_VEND', 'CURENCY_STATE', 'CURENCY_BASE', 'CURENCY_CONV', 'CURENCY_FACTOR', 'CURENCY_DATE', 'GL_ACCOUNT', 'COMPANYNO', 'DIVISION', 'DEPART', 'DOC_DATE', 'DUE_DATE', 'POSTED_AR_AP', 'PAYTO', 'PAYTO_SUBC', 'TERM_CODE', 'DOC_TOTAL', 'C_DOC_TOTAL', 'PAID_TOTAL', 'C_PAID_TOTAL', 'VOID_TOTAL', 'C_VOID_TOTAL', 'BALANCE_TOTAL', 'C_BALANCE_TOTAL', 'DISCOUNT_AMT', 'C_DISCOUNT_AMT', 'RECLOCK', 'ADDED_USR', 'ADDED_DTE', 'UPDATED_USR', 'UPDATED_DTE']
HDR_DEFAULTS = {'TableName': 'dbo.CHECK_HDR', 'DOC_TYPE': 'AR', 'DOC_CATEGORY': 'R2', 'DOC_STATUS': 1, 'CRDT_LINE': 0, 'MANUAL_AUTO': 'M', 'POST_GL': 'N', 'CLEARED': 'N', 'CLEARED_CURRENCY_STATE': 1, 'CURENCY_STATE': None, 'DISCOUNT_DISCREPANCY': 0, 'NETAMOUNT_DISCREPANCY': 0, 'CURRENCY_DISCREPANCY': 0, 'OPEN_BALANCE': 0, 'C_DISCOUNT_DISCREPANCY': 0, 'C_NETAMOUNT_DISCREPANCY': 0, 'C_CURRENCY_DISCREPANCY': 0, 'C_OPEN_BALANCE': 0, 'ORIG_DISCOUNT': 0, 'ORIG_DISCOUNT_DISCREPANCY': 0, 'ORIG_NETAMOUNT': 0, 'ORIG_NETAMOUNT_DISCREPANCY': 0, 'ORIG_AMOUNT': 0, 'ORIG_APPLIED': 0, 'ORIG_OPEN_BALANCE': 0, 'C_ORIG_AMOUNT': 0, 'C_ORIG_DISCOUNT': 0, 'C_ORIG_DISCOUNT_DISCREPANCY': 0, 'C_ORIG_NETAMOUNT': 0, 'C_ORIG_NETAMOUNT_DISCREPANCY': 0, 'C_ORIG_APPLIED': 0, 'C_ORIG_OPEN_BALANCE': 0, 'PAID_BY_ACH': 'N', 'ACCOUNT_ORIG_AMOUNT': 0, 'SIGN_1': 'N', 'SIGN_2': 'N', 'SIGN_3': 'N', 'SIGN_4': 'N', 'SIGN_5': 'N', 'SIGN_1_APR': 'N', 'SIGN_2_APR': 'N', 'SIGN_3_APR': 'N', 'SIGN_4_APR': 'N', 'SIGN_5_APR': 'N', 'RECLOCK': None}
# CHECK_LINE amount source rule used by generate():
#   Each line resolves its own invoice factor:
#     same-pair 1 -> INV_HDR factor -> ARAP factor -> TBLCONV fallback
#   C_AMOUNT     = INV_HDR.C_DOC_TOTAL
#   C_NETAMOUNT  = input C_PAID_AMOUNT (current receipt delta)
#   NETAMOUNT    = C_NETAMOUNT / line factor
#   C_DISCOUNT   = C_AMOUNT - C_NETAMOUNT
#
# CHECK_HDR factor rules:
#   ACCOUNT_FACTOR is sourced from the TBLCONV base-to-payment lookup;
#   ACCOUNT_AMOUNT / AMOUNT remains a validation relationship.
#   CURENCY_FACTOR uses the common line factor when every line has the
#   same factor. Otherwise it is the effective receipt factor
#   C_AMOUNT / AMOUNT.
#
# ACCOUNT_AR_AP paid totals are cumulative:
#   new C_PAID_TOTAL = existing C_PAID_TOTAL + current C_NETAMOUNT
#   new PAID_TOTAL   = existing PAID_TOTAL + current NETAMOUNT
#
# RMB and CNY are treated as the same currency.
# INV_LINE is not used by this generation path.
LINE_DEFAULTS = {'TableName': 'dbo.CHECK_LINE', 'DOC_TYPE': 'AR', 'DOC_CATEGORY': 'R2', 'DOC_STATUS': None, 'PAY_DOC_CATEGORY': 'IV', 'PAY_TYPE': 'I', 'DISCOUNT_DISCREPANCY': 0, 'NETAMOUNT_DISCREPANCY': 0, 'CURRENCY_DISCREPANCY': 0, 'ORIG_DISCOUNT': 0, 'ORIG_DISCOUNT_DISCREPANCY': 0, 'ORIG_NETAMOUNT': 0, 'ORIG_NETAMOUNT_DISCREPANCY': 0, 'ORIG_AMOUNT': 0, 'ORIG_CURRENCY_DISCREPANCY': 0, 'GL_CODE_AMT_DISCREPANCY': '790550-00', 'GL_CODE_DISC_DISCREPANCY_DB': '790550-00', 'GL_CODE_DISC_DISCREPANCY_CR': '790550-00', 'C_DISCOUNT_DISCREPANCY': 0, 'C_NETAMOUNT_DISCREPANCY': 0, 'C_CURRENCY_DISCREPANCY': 0, 'C_ORIG_DISCOUNT': 0, 'C_ORIG_DISCOUNT_DISCREPANCY': 0, 'C_ORIG_NETAMOUNT': 0, 'C_ORIG_NETAMOUNT_DISCREPANCY': 0, 'C_ORIG_AMOUNT': 0, 'C_ORIG_CURRENCY_DISCREPANCY': 0, 'CLEARED': 'N', 'RECLOCK': None}


# -----------------------------------------------------------------------------
# Generic helpers
# -----------------------------------------------------------------------------

def _text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def _currency(value: Any) -> str:
    code = _text(value).upper()
    return "CNY" if code == "RMB" else code


def _currency_db_aliases(value: Any) -> set[str]:
    """
    Return all database spellings that should be treated as one currency.

    P2000 data may use RMB and CNY interchangeably. Internal comparisons use
    CNY, while SQL retrieval includes both values so neither spelling is lost.
    """
    code = _currency(value)
    if code == "CNY":
        return {"CNY", "RMB"}
    return {code} if code else set()


def _number(value: Any, default: float | None = None) -> float | None:
    if value is None or (isinstance(value, float) and pd.isna(value)) or _text(value) == "":
        return default
    return float(value)


def _date(value: Any) -> pd.Timestamp | None:
    if value is None or (isinstance(value, float) and pd.isna(value)) or _text(value) == "":
        return None
    return pd.Timestamp(value)


def _money(value: Any) -> float | None:
    number = _number(value)
    return None if number is None else round(number, 2)


def _calculated_amount(value: Any) -> float | None:
    """Preserve useful precision for currency-converted database amounts."""
    number = _number(value)
    return None if number is None else round(number, 12)


def _user_doc(value: Any) -> str:
    """Normalize USER_DOC safely without fuzzy matching.

    - strip variable SQL/Excel outer padding;
    - remove only a numeric Excel-style trailing `.0`;
    - preserve existing leading zeroes;
    - left-pad numeric values to eight characters.
    """
    text = _text(value)
    if not text:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text.zfill(8) if text.isdigit() else text


def _status(actual: Any, expected: Any, tolerance: float = AMOUNT_TOLERANCE) -> str:
    a = _number(actual)
    e = _number(expected)
    if a is None or e is None:
        return "NOT_TESTED"
    return "MATCH" if abs(a - e) <= tolerance else "MISMATCH"


def _diff(actual: Any, expected: Any) -> float | None:
    a = _number(actual)
    e = _number(expected)
    return None if a is None or e is None else round(a - e, 10)


def _read_first_sheet(path: str | Path) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    return pd.read_excel(path, sheet_name=xls.sheet_names[0]).dropna(how="all").copy()


def _read_optional_sheet(
    path: str | Path,
    aliases: Iterable[str],
) -> pd.DataFrame:
    """Return an optional sheet, or an empty DataFrame when it is absent."""
    xls = pd.ExcelFile(path)
    by_upper = {name.strip().upper(): name for name in xls.sheet_names}
    for alias in aliases:
        actual = by_upper.get(str(alias).strip().upper())
        if actual:
            return pd.read_excel(path, sheet_name=actual).dropna(how="all").copy()
    return pd.DataFrame()


def _normalize_columns(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    result.columns = [str(column).strip().upper() for column in result.columns]
    return result


def _sql_values(values: Iterable[Any]) -> list[Any]:
    result: list[Any] = []
    seen: set[str] = set()
    for value in values:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            continue
        key = _text(value)
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def _read_sql_query(
    connection: Any,
    sql: str,
    params: Iterable[Any] = (),
) -> pd.DataFrame:
    """Read a SQL query through a pyodbc cursor without pandas' DBAPI warning."""
    cursor = connection.cursor()
    try:
        cursor.execute(sql, list(params))
        columns = [item[0] for item in cursor.description] if cursor.description else []
        rows = cursor.fetchall()
        return pd.DataFrame.from_records(rows, columns=columns)
    finally:
        cursor.close()


def _query_in(
    connection: Any,
    table: str,
    column: str,
    values: Iterable[Any],
    extra_where: str = "",
    extra_params: Iterable[Any] = (),
) -> pd.DataFrame:
    unique_values = _sql_values(values)
    if not unique_values:
        return pd.DataFrame()

    parts: list[pd.DataFrame] = []
    # Stay comfortably under SQL Server's 2,100-parameter limit.
    for start in range(0, len(unique_values), 900):
        batch = unique_values[start:start + 900]
        placeholders = ",".join("?" for _ in batch)
        sql = f"SELECT * FROM {table} WHERE {column} IN ({placeholders})"
        if extra_where:
            sql += f" AND ({extra_where})"
        params = list(batch) + list(extra_params)
        parts.append(_read_sql_query(connection, sql, params))

    return _normalize_columns(pd.concat(parts, ignore_index=True))


def _prepare_counter_rows(raw: pd.DataFrame) -> pd.DataFrame:
    if raw.empty:
        return raw

    required = {"DOC_CATEGORY", "DOC_TYPE", "COMPANYNO", "DIVISION", "COUNTER"}
    missing = required - set(raw.columns)
    if missing:
        raise ValueError(f"COUNTERSTBL is missing columns: {sorted(missing)}")

    rows: list[dict[str, Any]] = []
    for role, (doc_category, doc_type) in COUNTER_ROLE_MAP.items():
        matches = raw[
            raw["DOC_CATEGORY"].map(_text).eq(doc_category)
            & raw["DOC_TYPE"].map(_text).eq(doc_type)
        ]
        for _, row in matches.iterrows():
            record = row.to_dict()
            record["ROLE"] = role
            record["INCREMENT"] = 1
            record["COUNTER_SOURCE"] = "DB"
            rows.append(record)
    return pd.DataFrame(rows)


def _load_database_lookups(
    connection: Any,
    receipt_input: pd.DataFrame,
) -> tuple[
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
]:
    """Load references using PAY_USER_DOC as the user-facing invoice key.

    PAY_DOC_NO is deliberately not read from the input. ACCOUNT_AR_AP is first
    located by normalized USER_DOC + ACCTNO + DOC_CATEGORY='IV' + LINE=0; its
    DOC_NO is then used to load INV_HDR and later populate CHECK_LINE.PAY_DOC_NO.
    """
    user_docs = [
        _user_doc(value)
        for value in receipt_input["PAY_USER_DOC"].tolist()
        if _user_doc(value)
    ]
    bank_ids = receipt_input["BANK_ID"].tolist()
    acctnos = receipt_input["ACCTNO"].tolist()

    arap = _query_in(
        connection,
        "dbo.ACCOUNT_AR_AP",
        "RTRIM(USER_DOC)",
        user_docs,
        extra_where=(
            "ISNULL(LINE, 0) = 0 "
            "AND RTRIM(DOC_CATEGORY) = 'IV'"
        ),
    )

    arap_doc_nos = (
        pd.to_numeric(
            arap.get("DOC_NO", pd.Series(dtype=object)),
            errors="coerce",
        )
        .dropna()
        .astype(int)
        .tolist()
    )

    inv_hdr_table = os.getenv(
        "P2000_INV_HDR_TABLE",
        "dbo.INV_HDR",
    ).strip()
    inv_hdr = _query_in(
        connection,
        inv_hdr_table,
        "DOC_NO",
        arap_doc_nos,
    )

    banks = _query_in(connection, "dbo.BANKS_ACCOUNTS", "BANK_ID", bank_ids)
    customer_table = os.getenv(
        "P2000_CUSTOMER_TABLE",
        "dbo.CUSTVEND",
    ).strip()

    customers = _query_in(
        connection,
        customer_table,
        "ACCTNO",
        acctnos,
        extra_where="RTRIM(CUST_VEND) = ?",
        extra_params=["C"],
    )

    normalized_base_currencies = {
        _currency(value)
        for value in arap.get(
            "CURENCY_BASE",
            pd.Series(dtype=object),
        )
    }
    normalized_base_currencies.discard("")
    if not normalized_base_currencies:
        normalized_base_currencies = {"USD"}

    normalized_target_currencies = {
        _currency(value)
        for value in receipt_input["ACCOUNT_CURRENCY"]
    }
    normalized_target_currencies.update(
        _currency(value)
        for value in arap.get(
            "CURENCY_CONV",
            pd.Series(dtype=object),
        )
    )
    normalized_target_currencies.discard("")

    base_currencies = sorted(
        {
            alias
            for currency in normalized_base_currencies
            for alias in _currency_db_aliases(currency)
        }
    )
    target_currencies = sorted(
        {
            alias
            for currency in normalized_target_currencies
            for alias in _currency_db_aliases(currency)
        }
    )

    if target_currencies:
        fr_placeholders = ",".join("?" for _ in base_currencies)
        to_placeholders = ",".join("?" for _ in target_currencies)
        rate_sql = (
            "SELECT * FROM dbo.TBLCONV "
            f"WHERE UPPER(RTRIM(FR_CODE)) IN ({fr_placeholders}) "
            f"AND UPPER(RTRIM(TO_CODE)) IN ({to_placeholders})"
        )
        rates = _read_sql_query(
            connection,
            rate_sql,
            list(base_currencies) + list(target_currencies),
        )
        rates = _normalize_columns(rates)
    else:
        rates = pd.DataFrame()

    counter_sql = """
        SELECT *
        FROM dbo.COUNTERSTBL
        WHERE (RTRIM(DOC_CATEGORY) = 'CHECK' AND RTRIM(DOC_TYPE) = 'D')
           OR (RTRIM(DOC_CATEGORY) = 'R2'    AND RTRIM(DOC_TYPE) IN ('D', 'L'))
           OR (RTRIM(DOC_CATEGORY) = 'ACRN'  AND RTRIM(DOC_TYPE) = 'D')
           OR (RTRIM(DOC_CATEGORY) = 'BC'    AND RTRIM(DOC_TYPE) = 'D')
    """
    counters = _prepare_counter_rows(
        _normalize_columns(_read_sql_query(connection, counter_sql))
    )

    return arap, inv_hdr, banks, customers, rates, counters


def _read_alias(path: str | Path, aliases: list[str]) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    by_upper = {name.upper(): name for name in xls.sheet_names}
    for alias in aliases:
        actual = by_upper.get(alias.upper())
        if actual:
            return pd.read_excel(path, sheet_name=actual).dropna(how="all").copy()
    raise ValueError(f"Reference workbook is missing one of these sheets: {aliases}")


def _trim_table(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    if "TableName" in result.columns:
        mask = result["TableName"].map(_text).str.startswith("dbo.")
        if mask.any():
            result = result[mask].copy()
    return result.reset_index(drop=True)


def _init_row(columns: Iterable[str], defaults: dict[str, Any]) -> dict[str, Any]:
    row = {column: None for column in columns}
    for key, value in defaults.items():
        if key in row:
            row[key] = value
    return row


def _plain_header(ws, row_number: int = 1) -> None:
    for cell in ws[row_number]:
        if cell.value is None:
            continue
        font = copy(cell.font)
        font.bold = True
        cell.font = font
        fill = copy(cell.fill)
        fill.fill_type = "solid"
        fill.fgColor.rgb = "E7E6E6"
        cell.fill = fill


def _resize_plain(ws) -> None:
    for column_cells in ws.columns:
        values = [_text(cell.value) for cell in list(column_cells)[:200]]
        width = min(max(max((len(v) for v in values), default=8) + 2, 10), 28)
        ws.column_dimensions[column_cells[0].column_letter].width = width


def _format_text_columns(ws) -> None:
    text_headers = {
        "CHECK_NO", "PAY_USER_DOC", "USER_DOC", "ACCOUNT_NO",
        "GL_CODE", "GL_CODE_AMT", "GL_CODE_DISC", "TRANSFER_GL_CODE",
        "BANK_ID", "ACCTNO", "DOC_CATEGORY", "PAY_DOC_CATEGORY",
    }
    header_map = {cell.value: cell.column for cell in ws[1] if cell.value is not None}
    for header in text_headers:
        column = header_map.get(header)
        if column is None:
            continue
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=column).number_format = "@"


def _cell_state(value: Any) -> str:
    if value is None:
        return "NULL"
    try:
        if pd.isna(value):
            return "NULL"
    except (TypeError, ValueError):
        pass
    if isinstance(value, str) and value.strip() == "":
        return "BLANK"
    return "VALUE"


def _build_output_state_sheet(
    sheets: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    comparable_sheets = {
        "CHECK_HDR",
        "CHECK_LINE",
        "ARAP_UPDATE",
        "COUNTER_UPDATE",
    }

    for sheet_name, frame in sheets.items():
        if sheet_name not in comparable_sheets:
            continue
        for row_index, row in frame.reset_index(drop=True).iterrows():
            for column in frame.columns:
                state = _cell_state(row.get(column))
                if state in {"NULL", "BLANK"}:
                    rows.append(
                        {
                            "SHEET": sheet_name,
                            "OUTPUT_ROW_NUMBER": row_index + 2,
                            "COLUMN": column,
                            "VALUE_STATE": state,
                        }
                    )

    return pd.DataFrame(
        rows,
        columns=[
            "SHEET",
            "OUTPUT_ROW_NUMBER",
            "COLUMN",
            "VALUE_STATE",
        ],
    )


def _write_plain_sheets(path: str | Path, sheets: dict[str, pd.DataFrame]) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    sheets_to_write = dict(sheets)
    sheets_to_write["OUTPUT_STATE"] = _build_output_state_sheet(
        sheets_to_write
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets_to_write.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            _plain_header(ws, 1)
            _format_text_columns(ws)
            _resize_plain(ws)


def _write_validation_one_sheet(
    path: str | Path,
    header_validation: pd.DataFrame,
    line_validation: pd.DataFrame,
) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        header_validation.to_excel(writer, sheet_name="Validation", index=False, startrow=1)
        line_start = len(header_validation) + 5
        line_validation.to_excel(writer, sheet_name="Validation", index=False, startrow=line_start)
        ws = writer.book["Validation"]
        ws["A1"] = "HEADER VALIDATION"
        ws.cell(row=line_start, column=1, value="LINE VALIDATION")
        _plain_header(ws, 2)
        _plain_header(ws, line_start + 1)
        ws.freeze_panes = "A3"
        _resize_plain(ws)


# -----------------------------------------------------------------------------
# 1. Validate confirmed ARAP -> CHECK_LINE -> CHECK_HDR logic
# -----------------------------------------------------------------------------

def validate_existing_chain(
    arap_path: str | Path,
    check_line_path: str | Path,
    check_hdr_path: str | Path,
    output_path: str | Path,
    arap_history_line_path: str | Path | None = None,
    header_line_path: str | Path | None = None,
) -> None:
    """
    Validate the confirmed line-factor / effective-header-factor amount chain.

    CHECK_LINE uses an invoice-level factor. CHECK_HDR.CURENCY_FACTOR is an
    effective receipt-level factor calculated from the aggregated line amounts:

        line factor from amount = C_AMOUNT / AMOUNT
        line factor from net    = C_NETAMOUNT / NETAMOUNT

        header ACCOUNT_FACTOR   = ACCOUNT_AMOUNT / AMOUNT
        header CURENCY_FACTOR   = C_AMOUNT / AMOUNT

    ARAP PAID_TOTAL and C_PAID_TOTAL remain cumulative across all historical
    CHECK_LINE rows for the invoice.
    """
    arap = _normalize_columns(_trim_table(_read_first_sheet(arap_path)))
    line = _normalize_columns(_trim_table(_read_first_sheet(check_line_path)))
    hdr = _normalize_columns(_trim_table(_read_first_sheet(check_hdr_path)))
    header_line = (
        line.copy()
        if header_line_path is None
        else _normalize_columns(
            _trim_table(_read_first_sheet(header_line_path))
        )
    )
    arap_history_line = (
        line.copy()
        if arap_history_line_path is None
        else _normalize_columns(
            _trim_table(_read_first_sheet(arap_history_line_path))
        )
    )

    arap["USER_DOC_KEY"] = arap["USER_DOC"].map(_user_doc)
    arap["DOC_CATEGORY_KEY"] = arap["DOC_CATEGORY"].map(_text)
    arap["ACCTNO_KEY"] = arap["ACCTNO"].map(_text)
    arap = arap[arap.get("LINE", 0).fillna(0).eq(0)].copy()

    line["PAY_USER_DOC_KEY"] = line["PAY_USER_DOC"].map(_user_doc)
    line["PAY_DOC_CATEGORY_KEY"] = line["PAY_DOC_CATEGORY"].map(_text)
    line["ACCTNO_KEY"] = line["ACCTNO"].map(_text)
    line["DOC_CATEGORY_KEY"] = line["DOC_CATEGORY"].map(_text)

    header_line["PAY_DOC_CATEGORY_KEY"] = (
        header_line["PAY_DOC_CATEGORY"].map(_text)
    )
    header_line["DOC_CATEGORY_KEY"] = (
        header_line["DOC_CATEGORY"].map(_text)
    )

    arap_history_line["PAY_USER_DOC_KEY"] = (
        arap_history_line["PAY_USER_DOC"].map(_user_doc)
    )
    arap_history_line["PAY_DOC_CATEGORY_KEY"] = (
        arap_history_line["PAY_DOC_CATEGORY"].map(_text)
    )
    arap_history_line["ACCTNO_KEY"] = (
        arap_history_line["ACCTNO"].map(_text)
    )

    hdr["DOC_CATEGORY_KEY"] = hdr["DOC_CATEGORY"].map(_text)

    hdr_index: dict[tuple[Any, str], list[pd.Series]] = {}
    for _, row in hdr.iterrows():
        key = (
            _key_number(row.get("DOC_NO")),
            _text(row.get("DOC_CATEGORY")),
        )
        hdr_index.setdefault(key, []).append(row)

    arap_index: dict[tuple[Any, str, str, str], list[pd.Series]] = {}
    for _, row in arap.iterrows():
        key = (
            _key_number(row.get("DOC_NO")),
            _user_doc(row.get("USER_DOC")),
            _text(row.get("DOC_CATEGORY")),
            _text(row.get("ACCTNO")),
        )
        arap_index.setdefault(key, []).append(row)

    zero_default_columns = [
        "DISCOUNT_DISCREPANCY",
        "NETAMOUNT_DISCREPANCY",
        "CURRENCY_DISCREPANCY",
        "ORIG_DISCOUNT",
        "ORIG_DISCOUNT_DISCREPANCY",
        "ORIG_NETAMOUNT",
        "ORIG_NETAMOUNT_DISCREPANCY",
        "ORIG_AMOUNT",
        "ORIG_CURRENCY_DISCREPANCY",
        "C_DISCOUNT_DISCREPANCY",
        "C_NETAMOUNT_DISCREPANCY",
        "C_CURRENCY_DISCREPANCY",
        "C_ORIG_DISCOUNT",
        "C_ORIG_DISCOUNT_DISCREPANCY",
        "C_ORIG_NETAMOUNT",
        "C_ORIG_NETAMOUNT_DISCREPANCY",
        "C_ORIG_AMOUNT",
        "C_ORIG_CURRENCY_DISCREPANCY",
    ]

    line_rows: list[dict[str, Any]] = []

    for _, line_row in line.iterrows():
        receipt_key = (
            _key_number(line_row.get("DOC_NO")),
            _text(line_row.get("DOC_CATEGORY")),
        )
        invoice_key = (
            _key_number(line_row.get("PAY_DOC_NO")),
            _user_doc(line_row.get("PAY_USER_DOC")),
            _text(line_row.get("PAY_DOC_CATEGORY")),
            _text(line_row.get("ACCTNO")),
        )

        hdr_matches = hdr_index.get(receipt_key, [])
        arap_matches = arap_index.get(invoice_key, [])

        c_amount = _number(line_row.get("C_AMOUNT"))
        c_netamount = _number(line_row.get("C_NETAMOUNT"))
        c_discount = _number(line_row.get("C_DISCOUNT"))
        amount = _number(line_row.get("AMOUNT"))
        netamount = _number(line_row.get("NETAMOUNT"))
        discount = _number(line_row.get("DISCOUNT"))

        expected_c_discount = (
            None
            if c_amount is None or c_netamount is None
            else c_amount - c_netamount
        )
        expected_discount = (
            None
            if amount is None or netamount is None
            else amount - netamount
        )
        factor_from_amount = (
            None
            if c_amount is None or amount in (None, 0)
            else c_amount / amount
        )
        factor_from_netamount = (
            None
            if c_netamount is None or netamount in (None, 0)
            else c_netamount / netamount
        )

        formula_checks = {
            "C_DISCOUNT_EQ_C_AMOUNT_MINUS_C_NETAMOUNT":
                _status(c_discount, expected_c_discount),
            "DISCOUNT_EQ_AMOUNT_MINUS_NETAMOUNT":
                _status(discount, expected_discount),
            "LINE_FACTOR_AMOUNT_VS_NETAMOUNT":
                _status(
                    factor_from_amount,
                    factor_from_netamount,
                    tolerance=1e-6,
                ),
        }

        default_mismatches: list[str] = []
        for column in zero_default_columns:
            if column not in line.columns:
                continue
            value = _number(line_row.get(column))
            if value is not None and abs(value) > AMOUNT_TOLERANCE:
                default_mismatches.append(column)

        errors: list[str] = []
        warnings: list[str] = []
        if len(hdr_matches) == 0:
            errors.append("MISSING_CHECK_HDR")
        elif len(hdr_matches) > 1:
            errors.append("CHECK_HDR_NOT_UNIQUE")

        if len(arap_matches) == 0:
            errors.append("MISSING_ARAP")
        elif len(arap_matches) > 1:
            errors.append("ARAP_NOT_UNIQUE")

        failed_formula_checks = [
            name
            for name, status in formula_checks.items()
            if status == "MISMATCH"
        ]
        if failed_formula_checks:
            errors.append(
                "LINE_AMOUNT_MISMATCH:"
                + ",".join(failed_formula_checks)
            )

        if default_mismatches:
            warnings.append(
                "HISTORICAL_NONZERO_FIELDS:"
                + ",".join(default_mismatches)
            )

        line_rows.append(
            {
                "RECEIPT_DOC_NO": line_row.get("DOC_NO"),
                "LINE": line_row.get("LINE"),
                "PAY_DOC_NO": line_row.get("PAY_DOC_NO"),
                "PAY_USER_DOC": line_row.get("PAY_USER_DOC"),
                "ACCTNO": line_row.get("ACCTNO"),
                "HEADER_MATCH_COUNT": len(hdr_matches),
                "ARAP_MATCH_COUNT": len(arap_matches),
                "ACTUAL_C_AMOUNT": c_amount,
                "ACTUAL_AMOUNT": amount,
                "LINE_FACTOR_FROM_AMOUNT": factor_from_amount,
                "ACTUAL_C_NETAMOUNT": c_netamount,
                "ACTUAL_NETAMOUNT": netamount,
                "LINE_FACTOR_FROM_NETAMOUNT": factor_from_netamount,
                "LINE_FACTOR_DIFF": _diff(
                    factor_from_amount,
                    factor_from_netamount,
                ),
                "EXPECTED_C_DISCOUNT": expected_c_discount,
                "ACTUAL_C_DISCOUNT": c_discount,
                "C_DISCOUNT_DIFF": _diff(
                    c_discount,
                    expected_c_discount,
                ),
                "EXPECTED_DISCOUNT": expected_discount,
                "ACTUAL_DISCOUNT": discount,
                "DISCOUNT_DIFF": _diff(
                    discount,
                    expected_discount,
                ),
                "DEFAULT_ZERO_MISMATCHES":
                    ",".join(default_mismatches),
                "STATUS": "MATCH" if not errors else "MISMATCH",
                "ERRORS": " | ".join(errors),
                "WARNINGS": " | ".join(warnings),
            }
        )

    line_validation = pd.DataFrame(line_rows)

    line_totals = (
        header_line.groupby(
            ["DOC_NO", "DOC_CATEGORY_KEY"],
            dropna=False,
        )
        .agg(
            LINE_COUNT=("LINE", "count"),
            EXPECTED_APPLIED=("AMOUNT", "sum"),
            EXPECTED_AMOUNT=("NETAMOUNT", "sum"),
            EXPECTED_DISCOUNT=("DISCOUNT", "sum"),
            EXPECTED_C_APPLIED=("C_AMOUNT", "sum"),
            EXPECTED_C_AMOUNT=("C_NETAMOUNT", "sum"),
            EXPECTED_C_DISCOUNT=("C_DISCOUNT", "sum"),
        )
        .reset_index()
    )

    receipt_categories = (
        header_line.groupby(
            ["DOC_NO", "DOC_CATEGORY_KEY"],
            dropna=False,
        )["PAY_DOC_CATEGORY_KEY"]
        .agg(
            lambda values: ",".join(
                sorted(
                    {
                        _text(value) or "<BLANK>"
                        for value in values
                    }
                )
            )
        )
        .reset_index(name="PAY_DOC_CATEGORIES")
    )
    line_totals = line_totals.merge(
        receipt_categories,
        how="left",
        on=["DOC_NO", "DOC_CATEGORY_KEY"],
    )

    header = hdr.merge(
        line_totals,
        how="left",
        on=["DOC_NO", "DOC_CATEGORY_KEY"],
    )

    header_rows: list[dict[str, Any]] = []

    for _, row in header.iterrows():
        account_amount = _number(row.get("ACCOUNT_AMOUNT"))
        amount = _number(row.get("AMOUNT"))
        c_amount = _number(row.get("C_AMOUNT"))

        expected_account_factor = (
            None
            if account_amount is None or amount in (None, 0)
            else account_amount / amount
        )
        expected_currency_factor = (
            None
            if c_amount is None or amount in (None, 0)
            else c_amount / amount
        )

        checks = {
            "APPLIED_EQ_SUM_LINE_AMOUNT":
                _status(row.get("APPLIED"), row.get("EXPECTED_APPLIED")),
            "AMOUNT_EQ_SUM_LINE_NETAMOUNT":
                _status(row.get("AMOUNT"), row.get("EXPECTED_AMOUNT")),
            "NETAMOUNT_EQ_AMOUNT":
                _status(row.get("NETAMOUNT"), row.get("AMOUNT")),
            "DISCOUNT_EQ_SUM_LINE_DISCOUNT":
                _status(row.get("DISCOUNT"), row.get("EXPECTED_DISCOUNT")),
            "C_APPLIED_EQ_SUM_LINE_C_AMOUNT":
                _status(row.get("C_APPLIED"), row.get("EXPECTED_C_APPLIED")),
            "C_AMOUNT_EQ_SUM_LINE_C_NETAMOUNT":
                _status(row.get("C_AMOUNT"), row.get("EXPECTED_C_AMOUNT")),
            "C_NETAMOUNT_EQ_C_AMOUNT":
                _status(row.get("C_NETAMOUNT"), row.get("C_AMOUNT")),
            "C_DISCOUNT_EQ_SUM_LINE_C_DISCOUNT":
                _status(
                    row.get("C_DISCOUNT"),
                    row.get("EXPECTED_C_DISCOUNT"),
                ),
            "ACCOUNT_FACTOR_EQ_ACCOUNT_AMOUNT_DIV_AMOUNT":
                _status(
                    row.get("ACCOUNT_FACTOR"),
                    expected_account_factor,
                    tolerance=1e-6,
                ),
            "CURENCY_FACTOR_EQ_C_AMOUNT_DIV_AMOUNT":
                _status(
                    row.get("CURENCY_FACTOR"),
                    expected_currency_factor,
                    tolerance=1e-6,
                ),
        }

        failed_checks = [
            name
            for name, status in checks.items()
            if status == "MISMATCH"
        ]

        category_values = {
            value
            for value in _text(
                row.get("PAY_DOC_CATEGORIES")
            ).split(",")
            if value
        }
        non_iv_categories = sorted(
            value for value in category_values if value != "IV"
        )

        header_factor = _number(row.get("CURENCY_FACTOR"))
        applied_factor_residual = (
            None
            if header_factor in (None, 0)
            else _diff(
                row.get("APPLIED"),
                _number(row.get("C_APPLIED"), 0)
                / header_factor,
            )
        )
        amount_factor_residual = (
            None
            if header_factor in (None, 0)
            else _diff(
                row.get("AMOUNT"),
                _number(row.get("C_AMOUNT"), 0)
                / header_factor,
            )
        )

        factor_precision_checks = {
            "APPLIED_EQ_SUM_LINE_AMOUNT",
            "AMOUNT_EQ_SUM_LINE_NETAMOUNT",
        }
        factor_precision_only = (
            bool(failed_checks)
            and set(failed_checks).issubset(factor_precision_checks)
            and applied_factor_residual is not None
            and amount_factor_residual is not None
            and abs(applied_factor_residual) <= 0.000001
            and abs(amount_factor_residual) <= 0.000001
        )

        errors: list[str] = []
        warnings: list[str] = []
        if pd.isna(row.get("LINE_COUNT")):
            errors.append("NO_CHECK_LINE")

        if non_iv_categories:
            status = "OUT_OF_SCOPE"
            classification = (
                "OUT_OF_SCOPE_NON_IV_LINES:"
                + ",".join(non_iv_categories)
            )
            warnings.append(classification)
            if failed_checks:
                warnings.append(
                    "HISTORICAL_HEADER_DIFFERENCE:"
                    + ",".join(failed_checks)
                )
        elif factor_precision_only:
            status = "MATCH_WITHIN_FACTOR_PRECISION"
            classification = "FACTOR_PRECISION_ONLY"
            warnings.append(
                "HEADER_BASE_AMOUNT_FACTOR_PRECISION:"
                + ",".join(failed_checks)
            )
        elif failed_checks:
            status = "MISMATCH"
            classification = "UNEXPLAINED_CORE_MISMATCH"
            errors.append(
                "HDR_AMOUNT_MISMATCH:"
                + ",".join(failed_checks)
            )
        else:
            status = "MATCH"
            classification = "STANDARD_IV_MATCH"

        header_rows.append(
            {
                "DOC_NO": row.get("DOC_NO"),
                "CHECK_NO": row.get("CHECK_NO"),
                "LINE_COUNT": row.get("LINE_COUNT"),
                "EXPECTED_APPLIED": row.get("EXPECTED_APPLIED"),
                "ACTUAL_APPLIED": row.get("APPLIED"),
                "APPLIED_DIFF": _diff(
                    row.get("APPLIED"),
                    row.get("EXPECTED_APPLIED"),
                ),
                "EXPECTED_AMOUNT": row.get("EXPECTED_AMOUNT"),
                "ACTUAL_AMOUNT": row.get("AMOUNT"),
                "AMOUNT_DIFF": _diff(
                    row.get("AMOUNT"),
                    row.get("EXPECTED_AMOUNT"),
                ),
                "ACTUAL_NETAMOUNT": row.get("NETAMOUNT"),
                "NETAMOUNT_MINUS_AMOUNT": _diff(
                    row.get("NETAMOUNT"),
                    row.get("AMOUNT"),
                ),
                "EXPECTED_DISCOUNT": row.get("EXPECTED_DISCOUNT"),
                "ACTUAL_DISCOUNT": row.get("DISCOUNT"),
                "DISCOUNT_DIFF": _diff(
                    row.get("DISCOUNT"),
                    row.get("EXPECTED_DISCOUNT"),
                ),
                "EXPECTED_C_APPLIED": row.get("EXPECTED_C_APPLIED"),
                "ACTUAL_C_APPLIED": row.get("C_APPLIED"),
                "C_APPLIED_DIFF": _diff(
                    row.get("C_APPLIED"),
                    row.get("EXPECTED_C_APPLIED"),
                ),
                "EXPECTED_C_AMOUNT": row.get("EXPECTED_C_AMOUNT"),
                "ACTUAL_C_AMOUNT": row.get("C_AMOUNT"),
                "C_AMOUNT_DIFF": _diff(
                    row.get("C_AMOUNT"),
                    row.get("EXPECTED_C_AMOUNT"),
                ),
                "ACTUAL_C_NETAMOUNT": row.get("C_NETAMOUNT"),
                "C_NETAMOUNT_MINUS_C_AMOUNT": _diff(
                    row.get("C_NETAMOUNT"),
                    row.get("C_AMOUNT"),
                ),
                "EXPECTED_C_DISCOUNT": row.get("EXPECTED_C_DISCOUNT"),
                "ACTUAL_C_DISCOUNT": row.get("C_DISCOUNT"),
                "C_DISCOUNT_DIFF": _diff(
                    row.get("C_DISCOUNT"),
                    row.get("EXPECTED_C_DISCOUNT"),
                ),
                "EXPECTED_ACCOUNT_FACTOR": expected_account_factor,
                "ACTUAL_ACCOUNT_FACTOR": row.get("ACCOUNT_FACTOR"),
                "ACCOUNT_FACTOR_DIFF": _diff(
                    row.get("ACCOUNT_FACTOR"),
                    expected_account_factor,
                ),
                "EXPECTED_CURENCY_FACTOR": expected_currency_factor,
                "ACTUAL_CURENCY_FACTOR": row.get("CURENCY_FACTOR"),
                "CURENCY_FACTOR_DIFF": _diff(
                    row.get("CURENCY_FACTOR"),
                    expected_currency_factor,
                ),
                "PAY_DOC_CATEGORIES": row.get("PAY_DOC_CATEGORIES"),
                "NON_IV_CATEGORIES": ",".join(non_iv_categories),
                "APPLIED_FACTOR_RESIDUAL": applied_factor_residual,
                "AMOUNT_FACTOR_RESIDUAL": amount_factor_residual,
                "CLASSIFICATION": classification,
                "STATUS": status,
                "ERRORS": " | ".join(errors),
                "WARNINGS": " | ".join(warnings),
            }
        )

    header_validation = pd.DataFrame(header_rows)

    invoice_line_totals = (
        arap_history_line.groupby(
            [
                "PAY_DOC_NO",
                "PAY_USER_DOC_KEY",
                "PAY_DOC_CATEGORY_KEY",
                "ACCTNO_KEY",
            ],
            dropna=False,
        )
        .agg(
            PAYMENT_LINE_COUNT=("LINE", "count"),
            EXPECTED_PAID_TOTAL=("NETAMOUNT", "sum"),
            EXPECTED_C_PAID_TOTAL=("C_NETAMOUNT", "sum"),
        )
        .reset_index()
    )

    invoice_totals_index: dict[
        tuple[Any, str, str, str],
        pd.Series,
    ] = {}
    for _, row in invoice_line_totals.iterrows():
        key = (
            _key_number(row.get("PAY_DOC_NO")),
            _user_doc(row.get("PAY_USER_DOC_KEY")),
            _text(row.get("PAY_DOC_CATEGORY_KEY")),
            _text(row.get("ACCTNO_KEY")),
        )
        invoice_totals_index[key] = row

    arap_rows: list[dict[str, Any]] = []

    for _, arap_row in arap.iterrows():
        key = (
            _key_number(arap_row.get("DOC_NO")),
            _user_doc(arap_row.get("USER_DOC")),
            _text(arap_row.get("DOC_CATEGORY")),
            _text(arap_row.get("ACCTNO")),
        )
        totals = invoice_totals_index.get(key)

        if totals is None:
            arap_rows.append(
                {
                    "DOC_NO": arap_row.get("DOC_NO"),
                    "USER_DOC": arap_row.get("USER_DOC"),
                    "DOC_CATEGORY": arap_row.get("DOC_CATEGORY"),
                    "ACCTNO": arap_row.get("ACCTNO"),
                    "PAYMENT_LINE_COUNT": 0,
                    "EXPECTED_PAID_TOTAL": None,
                    "ACTUAL_PAID_TOTAL": arap_row.get("PAID_TOTAL"),
                    "PAID_TOTAL_DIFF": None,
                    "EXPECTED_C_PAID_TOTAL": None,
                    "ACTUAL_C_PAID_TOTAL": arap_row.get("C_PAID_TOTAL"),
                    "C_PAID_TOTAL_DIFF": None,
                    "STATUS": "NOT_TESTED",
                    "ERRORS": "",
                    "WARNINGS":
                        "NO_MATCHING_CHECK_LINE_IN_HISTORY_SCOPE",
                    "COVERAGE_NOTE": (
                        "Diagnostic only. ARAP totals may include reversals, "
                        "voids, or payment document categories outside the "
                        "confirmed AR/R2 generation scope."
                    ),
                }
            )
            continue

        expected_paid_total = totals.get("EXPECTED_PAID_TOTAL")
        expected_c_paid_total = totals.get("EXPECTED_C_PAID_TOTAL")

        actual_paid_total = _number(arap_row.get("PAID_TOTAL"))
        actual_c_paid_total = _number(
            arap_row.get("C_PAID_TOTAL")
        )
        paid_diff = (
            None
            if actual_paid_total is None
            or _number(expected_paid_total) is None
            else actual_paid_total - float(expected_paid_total)
        )
        c_paid_diff = (
            None
            if actual_c_paid_total is None
            or _number(expected_c_paid_total) is None
            else actual_c_paid_total - float(expected_c_paid_total)
        )
        arap_factor = abs(
            _number(arap_row.get("CURENCY_FACTOR"), 1.0) or 1.0
        )
        base_tolerance = AMOUNT_TOLERANCE
        converted_tolerance = max(
            AMOUNT_TOLERANCE,
            arap_factor * AMOUNT_TOLERANCE,
        )

        standard_base_match = (
            paid_diff is not None
            and abs(paid_diff) <= AMOUNT_TOLERANCE
        )
        standard_converted_match = (
            c_paid_diff is not None
            and abs(c_paid_diff) <= AMOUNT_TOLERANCE
        )
        scaled_rounding_match = (
            paid_diff is not None
            and c_paid_diff is not None
            and abs(paid_diff) <= base_tolerance
            and abs(c_paid_diff) <= converted_tolerance
        )

        if standard_base_match and standard_converted_match:
            status = "MATCH"
            classification = "EXACT_WITHIN_STANDARD_TOLERANCE"
            warnings: list[str] = []
        elif scaled_rounding_match:
            status = "MATCH_WITHIN_ROUNDING"
            classification = "FACTOR_SCALED_ROUNDING"
            warnings = [
                "ARAP_CUMULATIVE_ROUNDING_ONLY:"
                f"BASE_DIFF={round(paid_diff, 12)},"
                f"C_DIFF={round(c_paid_diff, 12)},"
                f"FACTOR={arap_factor}"
            ]
        else:
            failed_checks: list[str] = []
            if not standard_base_match:
                failed_checks.append(
                    "PAID_TOTAL_EQ_SUM_HISTORY_LINE_NETAMOUNT"
                )
            if not standard_converted_match:
                failed_checks.append(
                    "C_PAID_TOTAL_EQ_SUM_HISTORY_LINE_C_NETAMOUNT"
                )
            status = "REVIEW"
            classification = "UNCONFIRMED_CUMULATIVE_RULE"
            warnings = [
                "ARAP_CUMULATIVE_REVIEW:"
                + ",".join(failed_checks)
            ]

        arap_rows.append(
            {
                "DOC_NO": arap_row.get("DOC_NO"),
                "USER_DOC": arap_row.get("USER_DOC"),
                "DOC_CATEGORY": arap_row.get("DOC_CATEGORY"),
                "ACCTNO": arap_row.get("ACCTNO"),
                "PAYMENT_LINE_COUNT": totals.get("PAYMENT_LINE_COUNT"),
                "EXPECTED_PAID_TOTAL": expected_paid_total,
                "ACTUAL_PAID_TOTAL": arap_row.get("PAID_TOTAL"),
                "PAID_TOTAL_DIFF": (
                    None if paid_diff is None else round(paid_diff, 10)
                ),
                "EXPECTED_C_PAID_TOTAL": expected_c_paid_total,
                "ACTUAL_C_PAID_TOTAL": arap_row.get("C_PAID_TOTAL"),
                "C_PAID_TOTAL_DIFF": (
                    None if c_paid_diff is None else round(c_paid_diff, 10)
                ),
                "ARAP_CURENCY_FACTOR": arap_factor,
                "BASE_TOLERANCE": base_tolerance,
                "CONVERTED_TOLERANCE": converted_tolerance,
                "CLASSIFICATION": classification,
                "STATUS": status,
                "ERRORS": "",
                "WARNINGS": " | ".join(warnings),
                "COVERAGE_NOTE": (
                    "Diagnostic only. Exact equality is not used to gate "
                    "the confirmed CHECK_HDR/CHECK_LINE generation logic."
                ),
            }
        )

    arap_validation = pd.DataFrame(arap_rows)

    def _status_count(
        frame: pd.DataFrame,
        status: str,
    ) -> int:
        if frame.empty or "STATUS" not in frame.columns:
            return 0
        return int(frame["STATUS"].eq(status).sum())

    validation_summary = pd.DataFrame(
        [
            {
                "AREA": "CHECK_LINE_CORE",
                "GATING": "YES",
                "TOTAL": len(line_validation),
                "MATCH": _status_count(line_validation, "MATCH"),
                "MISMATCH": _status_count(
                    line_validation,
                    "MISMATCH",
                ),
                "REVIEW": 0,
                "NOT_TESTED": 0,
                "INTERPRETATION": (
                    "Formula and relationship validation. Historical "
                    "nonzero discrepancy/original fields are warnings."
                ),
            },
            {
                "AREA": "CHECK_HDR_CORE",
                "GATING": "YES",
                "TOTAL": len(header_validation),
                "MATCH": _status_count(header_validation, "MATCH"),
                "MISMATCH": _status_count(
                    header_validation,
                    "MISMATCH",
                ),
                "REVIEW": 0,
                "MATCH_WITHIN_FACTOR_PRECISION": _status_count(
                    header_validation,
                    "MATCH_WITHIN_FACTOR_PRECISION",
                ),
                "OUT_OF_SCOPE": _status_count(
                    header_validation,
                    "OUT_OF_SCOPE",
                ),
                "NOT_TESTED": 0,
                "INTERPRETATION": (
                    "Receipt totals use all CHECK_LINE categories. "
                    "AD/MS receipts are classified out of scope; "
                    "small base-side factor residuals are classified "
                    "as precision matches."
                ),
            },
            {
                "AREA": "ARAP_CUMULATIVE_DIAGNOSTIC",
                "GATING": "NO",
                "TOTAL": len(arap_validation),
                "MATCH": _status_count(arap_validation, "MATCH"),
                "MISMATCH": 0,
                "REVIEW": _status_count(arap_validation, "REVIEW"),
                "MATCH_WITHIN_ROUNDING": _status_count(
                    arap_validation,
                    "MATCH_WITHIN_ROUNDING",
                ),
                "NOT_TESTED": _status_count(
                    arap_validation,
                    "NOT_TESTED",
                ),
                "INTERPRETATION": (
                    "Diagnostic only until the full inclusion rules for "
                    "voids, reversals, and other payment categories are "
                    "confirmed."
                ),
            },
        ]
    )

    _write_plain_sheets(
        output_path,
        {
            "Validation_Summary": validation_summary,
            "Header_Validation": header_validation,
            "Line_Validation": line_validation,
            "ARAP_Diagnostic": arap_validation,
        },
    )


# -----------------------------------------------------------------------------
# 2. Generate full Excel rows from the simple one-sheet input
# -----------------------------------------------------------------------------

def _lookup_rate(
    rates: pd.DataFrame,
    base_currency: str,
    target_currency: str,
) -> tuple[float | None, pd.Timestamp | None, str]:
    base = _currency(base_currency)
    target = _currency(target_currency)
    if not base or not target:
        return None, None, "INVALID_CURRENCY"
    if base == target:
        return 1.0, None, "SAME"

    work = rates.copy()
    work["FR_CODE_N"] = work["FR_CODE"].map(_currency)
    work["TO_CODE_N"] = work["TO_CODE"].map(_currency)
    work["FACTOR_N"] = pd.to_numeric(work["FACTOR"], errors="coerce")
    work["ADDED_DTE_N"] = pd.to_datetime(work.get("ADDED_DTE"), errors="coerce")
    work["CONVDATE_N"] = pd.to_datetime(work.get("CONVDATE"), errors="coerce")
    work = work[
        (work["FR_CODE_N"] == base)
        & (work["TO_CODE_N"] == target)
        & work["FACTOR_N"].notna()
        & work["FACTOR_N"].ne(0)
    ].copy()
    if work.empty:
        return None, None, "NO_RATE"

    sort_cols = [c for c in ["ADDED_DTE_N", "CONVDATE_N", "TBLCONV_DOC_NO"] if c in work.columns]
    work = work.sort_values(sort_cols, ascending=False, na_position="last")
    row = work.iloc[0]
    factor_date = row.get("CONVDATE_N")
    if pd.isna(factor_date):
        factor_date = row.get("ADDED_DTE_N")
    return float(row["FACTOR_N"]), factor_date, "LOOKUP"


def _apply_counter_mock_overrides(
    counters: pd.DataFrame,
    mock_counters: pd.DataFrame,
) -> pd.DataFrame:
    """
    Apply optional Excel counter snapshots.

    MockCounters columns:
        ROLE, COMPANYNO, DIVISION, COUNTER_BEFORE

    A nonblank COUNTER_BEFORE overrides the database counter. A blank value
    means "use the database lookup". The supplied number is the last-used
    counter value, so allocation still starts at COUNTER_BEFORE + 1.
    """
    if mock_counters.empty:
        return counters

    mock = _normalize_columns(mock_counters)
    required = {"ROLE", "COMPANYNO", "DIVISION", "COUNTER_BEFORE"}
    missing = required - set(mock.columns)
    if missing:
        raise ValueError(
            "MockCounters sheet is missing columns: " + ", ".join(sorted(missing))
        )

    result = counters.copy()
    if "COUNTER_SOURCE" not in result.columns:
        result["COUNTER_SOURCE"] = "DB"

    for _, row in mock.iterrows():
        role = _text(row.get("ROLE")).upper()
        before = _number(row.get("COUNTER_BEFORE"))
        if not role or before is None:
            # Blank COUNTER_BEFORE means use DB.
            continue
        if role not in COUNTER_ROLE_MAP:
            raise ValueError(f"MockCounters contains unsupported ROLE={role}")

        companyno = row.get("COMPANYNO")
        division = row.get("DIVISION")
        company_key = CounterAllocator._key_part(companyno)
        division_key = CounterAllocator._key_part(division)

        if result.empty:
            exact_mask = pd.Series(dtype=bool)
        else:
            exact_mask = result.apply(
                lambda current: (
                    _text(current.get("ROLE")).upper() == role
                    and CounterAllocator._key_part(current.get("COMPANYNO")) == company_key
                    and CounterAllocator._key_part(current.get("DIVISION")) == division_key
                ),
                axis=1,
            )

        matched_indexes = result.index[exact_mask].tolist() if len(exact_mask) else []

        # Permit the same unique NULL/BLANK fallback used by CounterAllocator.
        if not matched_indexes and division_key in {"<NULL>", "<BLANK>"} and not result.empty:
            fallback_mask = result.apply(
                lambda current: (
                    _text(current.get("ROLE")).upper() == role
                    and CounterAllocator._key_part(current.get("COMPANYNO")) == company_key
                    and CounterAllocator._key_part(current.get("DIVISION"))
                        in {"<NULL>", "<BLANK>"}
                ),
                axis=1,
            )
            matched_indexes = result.index[fallback_mask].tolist()

        if matched_indexes:
            # Update every duplicate physical DB row for the same logical key.
            result.loc[matched_indexes, "COUNTER"] = int(before)
            result.loc[matched_indexes, "COUNTER_SOURCE"] = "EXCEL_MOCK"
            continue

        doc_category, doc_type = COUNTER_ROLE_MAP[role]
        new_row = {
            "ROLE": role,
            "DOC_CATEGORY": doc_category,
            "DOC_TYPE": doc_type,
            "COMPANYNO": companyno,
            "DIVISION": division,
            "COUNTER": int(before),
            "INCREMENT": 1,
            "COUNTER_SOURCE": "EXCEL_MOCK",
        }
        result = pd.concat([result, pd.DataFrame([new_row])], ignore_index=True)

    return result


@dataclass
class CounterAllocator:
    frame: pd.DataFrame

    @staticmethod
    def _key_part(value: Any) -> str:
        """
        Preserve SQL NULL and blank text as different counter-key values.

        COUNTERSTBL may legitimately contain both:
            DIVISION IS NULL
            DIVISION = '' / spaces

        They must not collapse into the same key.
        """
        if value is None:
            return "<NULL>"

        try:
            if pd.isna(value):
                return "<NULL>"
        except (TypeError, ValueError):
            pass

        if isinstance(value, str):
            stripped = value.strip()
            return stripped if stripped else "<BLANK>"

        number = _number(value)
        if number is not None and float(number).is_integer():
            return str(int(number))

        text_value = _text(value)
        return text_value if text_value else "<BLANK>"

    def __post_init__(self) -> None:
        self.frame = self.frame.copy()
        required = {"ROLE", "COMPANYNO", "DIVISION", "COUNTER"}
        missing = required - set(self.frame.columns)
        if missing:
            raise ValueError(f"COUNTERSTBL lookup is missing columns: {sorted(missing)}")

        self.frame["ROLE"] = self.frame["ROLE"].map(_text)
        self.frame["COUNTER_KEY"] = self.frame.apply(
            lambda row: (
                _text(row.get("ROLE")),
                self._key_part(row.get("COMPANYNO")),
                self._key_part(row.get("DIVISION")),
            ),
            axis=1,
        )
        # COUNTERSTBL can contain duplicate physical rows for the same logical
        # counter key. If every duplicate carries the same COUNTER value, treat
        # them as one logical counter for read-only simulation. If their values
        # differ, stop because selecting one would be unsafe.
        if self.frame["COUNTER_KEY"].duplicated().any():
            collapsed_rows: list[pd.Series] = []
            conflicting_groups: list[pd.DataFrame] = []

            for _, group in self.frame.groupby("COUNTER_KEY", sort=False, dropna=False):
                counter_values = {
                    int(_number(value, 0) or 0)
                    for value in group["COUNTER"].tolist()
                }
                increment_values = {
                    int(_number(value, 1) or 1)
                    for value in group.get(
                        "INCREMENT",
                        pd.Series([1] * len(group), index=group.index),
                    ).tolist()
                }

                if len(counter_values) > 1 or len(increment_values) > 1:
                    conflicting_groups.append(group.copy())
                    continue

                row = group.iloc[0].copy()
                row["SOURCE_ROW_COUNT"] = len(group)
                collapsed_rows.append(row)

            if conflicting_groups:
                conflict = pd.concat(conflicting_groups, ignore_index=True)
                columns = [
                    column
                    for column in [
                        "ROLE",
                        "DOC_CATEGORY",
                        "DOC_TYPE",
                        "COMPANYNO",
                        "DIVISION",
                        "COUNTER",
                        "INCREMENT",
                    ]
                    if column in conflict.columns
                ]
                raise ValueError(
                    "COUNTERSTBL contains duplicate logical keys with different "
                    "COUNTER or INCREMENT values:\n"
                    + conflict[columns].to_string(index=False)
                )

            self.frame = pd.DataFrame(collapsed_rows).reset_index(drop=True)
        else:
            self.frame["SOURCE_ROW_COUNT"] = 1

        self.before = {
            row["COUNTER_KEY"]: int(_number(row.get("COUNTER"), 0) or 0)
            for _, row in self.frame.iterrows()
        }
        self.current = self.before.copy()
        self.increment = {
            row["COUNTER_KEY"]: int(_number(row.get("INCREMENT"), 1) or 1)
            for _, row in self.frame.iterrows()
        }

    def _resolve_key(
        self,
        role: str,
        companyno: Any,
        division: Any,
    ) -> tuple[str, str, str]:
        role_key = _text(role)
        company_key = self._key_part(companyno)
        division_key = self._key_part(division)

        exact_key = (role_key, company_key, division_key)
        if exact_key in self.current:
            return exact_key

        # SQL NULL and fixed-width blank text are distinct in storage, but many
        # P2000 counter rows use one form while the source business row uses the
        # other. Only fall back between these two forms when exactly one logical
        # candidate exists.
        if division_key in {"<NULL>", "<BLANK>"}:
            candidates = [
                key
                for key in self.current
                if key[0] == role_key
                and key[1] == company_key
                and key[2] in {"<NULL>", "<BLANK>"}
            ]

            if len(candidates) == 1:
                return candidates[0]

            if len(candidates) > 1:
                details = [
                    {
                        "DIVISION_KEY": key[2],
                        "COUNTER": self.current[key],
                        "INCREMENT": self.increment[key],
                    }
                    for key in candidates
                ]
                raise ValueError(
                    "COUNTERSTBL has ambiguous NULL/BLANK rows for "
                    f"ROLE={role}, COMPANYNO={companyno}: {details}"
                )

        available = [
            key[2]
            for key in self.current
            if key[0] == role_key and key[1] == company_key
        ]
        raise ValueError(
            "COUNTERSTBL has no unique row for "
            f"ROLE={role}, COMPANYNO={companyno}, DIVISION={division!r}. "
            f"Available division keys: {available}"
        )

    def next(self, role: str, companyno: Any, division: Any) -> int:
        key = self._resolve_key(role, companyno, division)
        self.current[key] += self.increment[key]
        return self.current[key]

    def output(self) -> pd.DataFrame:
        result = self.frame.copy()
        result["COUNTER_BEFORE"] = result["COUNTER_KEY"].map(self.before)
        result["COUNTER_AFTER"] = result["COUNTER_KEY"].map(self.current)
        result["INCREMENT_USED"] = (
            result["COUNTER_AFTER"] - result["COUNTER_BEFORE"]
        )
        result = result[result["INCREMENT_USED"].ne(0)].copy()
        result["STATUS"] = "SIMULATED_UPDATE"

        columns = [
            "ROLE",
            "DOC_CATEGORY",
            "DOC_TYPE",
            "COMPANYNO",
            "DIVISION",
            "COUNTER_SOURCE",
            "COUNTER_BEFORE",
            "COUNTER_AFTER",
            "INCREMENT_USED",
            "SOURCE_ROW_COUNT",
            "STATUS",
        ]
        return result[
            [column for column in columns if column in result.columns]
        ].reset_index(drop=True)


def _single(group: pd.DataFrame, column: str, errors: list[str], required: bool = True) -> Any:
    if column not in group.columns:
        if required:
            errors.append(f"MISSING_{column}")
        return None
    values = [v for v in group[column].tolist() if not pd.isna(v) and _text(v) != ""]
    normalized = list(dict.fromkeys(_text(v) if isinstance(v, str) else v for v in values))
    if len(normalized) > 1:
        errors.append(f"INCONSISTENT_{column}")
    if not normalized:
        if required:
            errors.append(f"MISSING_{column}")
        return None
    return normalized[0]


def _valid_factor(value: Any) -> float | None:
    factor = _number(value)
    return factor if factor not in (None, 0) else None


def _resolve_line_currency_factor(
    source: pd.Series,
    rates: pd.DataFrame,
    default_base_currency: str,
) -> tuple[
    float | None,
    pd.Timestamp | None,
    str,
    str,
    str,
]:
    """Resolve one invoice line's own currency factor."""
    line_base = (
        _currency(source.get("INV_CURENCY_BASE"))
        or _currency(source.get("CURENCY_BASE"))
        or _currency(default_base_currency)
        or "USD"
    )
    line_conv = (
        _currency(source.get("INV_CURENCY_CONV"))
        or _currency(source.get("CURENCY_CONV"))
        or line_base
    )

    inv_date = _date(source.get("INV_CURENCY_DATE"))
    arap_date = _date(source.get("CURENCY_DATE"))

    if line_base == line_conv:
        return (
            1.0,
            inv_date or arap_date,
            "SAME_CURRENCY_PAIR",
            line_base,
            line_conv,
        )

    inv_factor = _valid_factor(source.get("INV_CURENCY_FACTOR"))
    if inv_factor is not None:
        return (
            inv_factor,
            inv_date,
            "INV_HDR",
            line_base,
            line_conv,
        )

    arap_factor = _valid_factor(source.get("CURENCY_FACTOR"))
    if arap_factor is not None:
        return (
            arap_factor,
            arap_date,
            "ARAP",
            line_base,
            line_conv,
        )

    factor, factor_date, rate_status = _lookup_rate(
        rates,
        line_base,
        line_conv,
    )
    return (
        factor,
        factor_date,
        "TBLCONV_FALLBACK:" + rate_status,
        line_base,
        line_conv,
    )


def _max_currency_state(group: pd.DataFrame) -> tuple[Any, str]:
    candidates: list[float] = []
    sources: list[str] = []

    for column, source_name in [
        ("INV_CURENCY_STATE", "INV_HDR"),
        ("CURENCY_STATE", "ARAP"),
    ]:
        if column not in group.columns:
            continue
        column_values: list[float] = []
        for value in group[column].tolist():
            number = _number(value)
            if number is not None:
                column_values.append(number)
        if column_values:
            candidates.extend(column_values)
            sources.append(source_name)

    if not candidates:
        return 1, "DEFAULT_1"

    state = max(candidates)
    if float(state).is_integer():
        state = int(state)
    return state, "MAX_" + "+".join(sources)


def generate_insert_ready_excel(
    input_path: str | Path,
    output_path: str | Path,
    connection_string: str,
    user: str = "ILC",
    process_time: str | pd.Timestamp | None = None,
) -> None:
    process_ts = (
        pd.Timestamp(process_time)
        if process_time
        else pd.Timestamp.now().floor("s")
    )

    receipt_input = _normalize_columns(_read_first_sheet(input_path))

    # Canonical v22 names. Accept the old grouping label only as a convenience;
    # it is never inserted into CHECK_HDR.DOC_NO.
    if (
        "INPUT_HDR_GROUP_ID" not in receipt_input.columns
        and "RECEIPT_KEY" in receipt_input.columns
    ):
        receipt_input = receipt_input.rename(
            columns={"RECEIPT_KEY": "INPUT_HDR_GROUP_ID"}
        )
    if (
        "C_PAID_AMOUNT" not in receipt_input.columns
        and "C_PAID_TOTAL" in receipt_input.columns
    ):
        receipt_input = receipt_input.rename(
            columns={"C_PAID_TOTAL": "C_PAID_AMOUNT"}
        )

    required = {
        "INPUT_HDR_GROUP_ID",
        "BANK_ID",
        "ACCOUNT_CURRENCY",
        "C_PAID_AMOUNT",
        "CHECK_DATE",
        "PAY_USER_DOC",
        "ACCTNO",
        "GL_CODE_DISC",
    }
    missing = sorted(required - set(receipt_input.columns))
    if missing:
        raise ValueError(f"Input sheet is missing columns: {missing}")

    if "NOTE" not in receipt_input.columns:
        receipt_input["NOTE"] = None

    receipt_input = receipt_input.dropna(
        subset=["INPUT_HDR_GROUP_ID", "PAY_USER_DOC", "ACCTNO"]
    ).copy()
    receipt_input["INPUT_HDR_GROUP_ID"] = receipt_input[
        "INPUT_HDR_GROUP_ID"
    ].map(_text)
    receipt_input["PAY_USER_DOC_KEY"] = receipt_input["PAY_USER_DOC"].map(
        _user_doc
    )
    receipt_input["PAY_DOC_CATEGORY_KEY"] = "IV"
    receipt_input["ACCTNO_KEY"] = receipt_input["ACCTNO"].map(_text)

    mock_counters = _read_optional_sheet(
        input_path,
        ["MockCounters", "Mock_Counters", "CounterMock"],
    )

    try:
        import pyodbc
    except ImportError as exc:
        raise RuntimeError(
            "pyodbc is required for database lookups: pip install pyodbc"
        ) from exc

    if not connection_string:
        raise ValueError(
            "SQL Server connection settings are missing. Configure the "
            "P2000_DB_* variables in .env or in the operating-system "
            "environment."
        )

    with pyodbc.connect(connection_string) as connection:
        (
            arap,
            inv_hdr,
            banks,
            customers,
            rates,
            counters,
        ) = _load_database_lookups(connection, receipt_input)

    arap = _trim_table(arap)
    inv_hdr = _trim_table(inv_hdr)
    counters = _apply_counter_mock_overrides(counters, mock_counters)

    arap["USER_DOC_KEY"] = arap["USER_DOC"].map(_user_doc)
    arap["DOC_CATEGORY_KEY"] = arap["DOC_CATEGORY"].map(_text)
    arap["ACCTNO_KEY"] = arap["ACCTNO"].map(_text)
    arap = arap[
        arap.get("LINE", 0).fillna(0).eq(0)
        & arap["DOC_CATEGORY_KEY"].eq("IV")
    ].copy()

    inv_required = {"DOC_NO", "DOC_TOTAL", "C_DOC_TOTAL"}
    inv_missing = sorted(inv_required - set(inv_hdr.columns))
    if inv_missing:
        raise ValueError(
            "INV_HDR lookup is missing columns: "
            + ", ".join(inv_missing)
        )

    inv_optional = [
        "CURENCY_STATE",
        "CURENCY_BASE",
        "CURENCY_CONV",
        "CURENCY_FACTOR",
        "CURENCY_DATE",
    ]
    inv_columns = ["DOC_NO", "DOC_TOTAL", "C_DOC_TOTAL"] + [
        column
        for column in inv_optional
        if column in inv_hdr.columns
    ]
    inv_details = inv_hdr[inv_columns].copy()

    if inv_details["DOC_NO"].duplicated().any():
        duplicates = (
            inv_details.loc[
                inv_details["DOC_NO"].duplicated(False),
                "DOC_NO",
            ]
            .drop_duplicates()
            .tolist()
        )
        raise ValueError(
            "INV_HDR contains duplicate DOC_NO rows for input invoices: "
            + ", ".join(str(value) for value in duplicates)
        )

    inv_rename = {
        "DOC_TOTAL": "INV_DOC_TOTAL",
        "C_DOC_TOTAL": "INV_C_DOC_TOTAL",
        "CURENCY_STATE": "INV_CURENCY_STATE",
        "CURENCY_BASE": "INV_CURENCY_BASE",
        "CURENCY_CONV": "INV_CURENCY_CONV",
        "CURENCY_FACTOR": "INV_CURENCY_FACTOR",
        "CURENCY_DATE": "INV_CURENCY_DATE",
    }
    inv_details = inv_details.rename(columns=inv_rename)

    arap = arap.merge(
        inv_details,
        how="left",
        on="DOC_NO",
        validate="many_to_one",
        indicator="INV_HDR_MERGE",
    )

    try:
        merged = receipt_input.merge(
            arap,
            how="left",
            left_on=[
                "PAY_USER_DOC_KEY",
                "PAY_DOC_CATEGORY_KEY",
                "ACCTNO_KEY",
            ],
            right_on=[
                "USER_DOC_KEY",
                "DOC_CATEGORY_KEY",
                "ACCTNO_KEY",
            ],
            suffixes=("_INPUT", "_ARAP"),
            indicator=True,
            validate="many_to_one",
        )
    except pd.errors.MergeError as exc:
        raise ValueError(
            "ACCOUNT_AR_AP is not unique for normalized "
            "PAY_USER_DOC + ACCTNO + DOC_CATEGORY='IV' + LINE=0. "
            "PAY_DOC_NO cannot be derived safely."
        ) from exc

    allocator = CounterAllocator(counters)
    hdr_rows: list[dict[str, Any]] = []
    line_rows: list[dict[str, Any]] = []
    arap_rows: list[dict[str, Any]] = []
    validation_rows: list[dict[str, Any]] = []
    batch_numbers: dict[tuple[str, str], int] = {}

    for input_hdr_group_id, group in merged.groupby("INPUT_HDR_GROUP_ID", sort=False):
        errors: list[str] = []
        if (group["_merge"] != "both").any():
            errors.append("MISSING_ARAP")

        matched_arap = group[group["_merge"] == "both"]
        if (
            not matched_arap.empty
            and matched_arap["INV_HDR_MERGE"].ne("both").any()
        ):
            errors.append("MISSING_INV_HDR")

        bank_id = _single(group, "BANK_ID", errors)
        account_currency = _currency(
            _single(group, "ACCOUNT_CURRENCY", errors)
        )

        check_date = _date(_single(group, "CHECK_DATE", errors)) or process_ts
        acctno = _text(_single(group, "ACCTNO_INPUT", errors))
        note_value = _single(group, "NOTE", errors, required=False)
        note_value = _text(note_value) or None

        matched = group[group["_merge"] == "both"].copy()
        arap_base_values = sorted(
            {
                _currency(value)
                for value in matched.get(
                    "CURENCY_BASE",
                    pd.Series(dtype=object),
                )
                if _currency(value)
            }
        )
        default_base_currency = (
            arap_base_values[0]
            if arap_base_values
            else "USD"
        )

        subc_values = [
            value
            for value in matched.get(
                "SUBC",
                pd.Series(dtype=object),
            ).tolist()
            if not pd.isna(value)
        ]
        header_subc = subc_values[0] if subc_values else None
        if len({_text(value) for value in subc_values}) > 1:
            errors.append("MIXED_SUBC")

        bank_id_matches = banks[
            banks["BANK_ID"].map(_text) == _text(bank_id)
        ]
        bank_currency_column = "CURENCY_BASE"

        if bank_currency_column not in banks.columns:
            bank_matches = bank_id_matches.iloc[0:0]
            errors.append("BANKS_ACCOUNTS_MISSING_CURENCY_BASE_COLUMN")
        else:
            bank_matches = bank_id_matches[
                bank_id_matches[bank_currency_column]
                .map(_currency)
                .eq(account_currency)
            ]

        bank = (
            bank_matches.iloc[0]
            if len(bank_matches) == 1
            else pd.Series(dtype=object)
        )
        if len(bank_matches) != 1:
            errors.append(
                "BANK_LOOKUP_NOT_UNIQUE:"
                f"BANK_ID={_text(bank_id)},"
                f"CURENCY_BASE={account_currency},"
                f"MATCH_COUNT={len(bank_matches)}"
            )

        customer_matches = customers[
            (customers["CUST_VEND"].map(_text) == "C")
            & (customers["ACCTNO"].map(_text) == acctno)
        ]

        if "SUBC" in customers.columns:
            customer_matches = customer_matches[
                customer_matches["SUBC"]
                .map(_text)
                .eq(_text(header_subc))
            ]
        else:
            errors.append("CUSTVEND_MISSING_SUBC_COLUMN")
            customer_matches = customer_matches.iloc[0:0]

        unique_customer_names = sorted(
            {
                _text(value)
                for value in customer_matches.get(
                    "NAME",
                    pd.Series(dtype=object),
                ).tolist()
                if _text(value)
            }
        )

        if len(unique_customer_names) == 1:
            customer = customer_matches.iloc[0].copy()
            customer["NAME"] = unique_customer_names[0]
            customer_lookup_status = (
                "UNIQUE_ROW"
                if len(customer_matches) == 1
                else "DUPLICATE_ROWS_SAME_NAME"
            )
        else:
            customer = pd.Series(dtype=object)
            customer_lookup_status = (
                "NO_MATCH"
                if len(customer_matches) == 0
                else "MULTIPLE_NAMES"
            )
            customer_error_name = (
                "CUSTOMER_LOOKUP_NO_MATCH"
                if len(customer_matches) == 0
                else "CUSTOMER_LOOKUP_NOT_UNIQUE"
            )
            errors.append(
                customer_error_name + ":"
                f"ACCTNO={acctno},"
                f"SUBC={_text(header_subc)},"
                f"MATCH_COUNT={len(customer_matches)},"
                f"UNIQUE_NAMES={unique_customer_names}"
            )

        (
            payment_lookup_factor,
            payment_lookup_factor_date,
            payment_lookup_status,
        ) = _lookup_rate(
            rates,
            default_base_currency,
            account_currency,
        )

        hdr_currency_state, currency_state_source = _max_currency_state(
            matched
        )

        companyno_values = [
            value
            for value in matched.get(
                "COMPANYNO",
                pd.Series(dtype=object),
            ).tolist()
            if not pd.isna(value)
        ]
        companyno = companyno_values[0] if companyno_values else None
        if len(
            {
                CounterAllocator._key_part(value)
                for value in companyno_values
            }
        ) > 1:
            errors.append("MIXED_COMPANYNO")

        counter_division = bank.get("DIVISION")
        if counter_division is None or _text(counter_division) == "":
            division_values = [
                value
                for value in matched.get(
                    "DIVISION",
                    pd.Series(dtype=object),
                ).tolist()
                if not pd.isna(value)
            ]
            counter_division = (
                division_values[0]
                if division_values
                else None
            )

        try:
            batch_key = (
                CounterAllocator._key_part(companyno),
                CounterAllocator._key_part(counter_division),
            )
            if batch_key not in batch_numbers:
                batch_numbers[batch_key] = allocator.next(
                    "BC_BATCH_NO",
                    companyno,
                    counter_division,
                )

            batch_no = batch_numbers[batch_key]
            doc_no = allocator.next(
                "CHECK_DOC_NO",
                companyno,
                counter_division,
            )
            check_no = str(
                allocator.next(
                    "R2_CHECK_NO",
                    companyno,
                    counter_division,
                )
            ).zfill(6)
            check_counter = allocator.next(
                "ACRN_CHECK_COUNTER",
                companyno,
                counter_division,
            )
        except ValueError as exc:
            errors.append(str(exc))
            batch_no = None
            doc_no = None
            check_no = None
            check_counter = None

        calculated: list[dict[str, Any]] = []

        for _, source in group.iterrows():
            if source.get("_merge") != "both":
                continue

            pay_doc_no = int(source["DOC_NO"])
            current_c_paid_amount = _calculated_amount(
                source.get("C_PAID_AMOUNT")
            )
            inv_doc_total = _calculated_amount(
                source.get("INV_DOC_TOTAL")
            )
            inv_c_doc_total = _calculated_amount(
                source.get("INV_C_DOC_TOTAL")
            )

            if source.get("INV_HDR_MERGE") != "both":
                errors.append(f"MISSING_INV_HDR_{pay_doc_no}")
                continue
            if inv_doc_total is None:
                errors.append(f"MISSING_INV_DOC_TOTAL_{pay_doc_no}")
                continue
            if inv_c_doc_total is None:
                errors.append(f"MISSING_INV_C_DOC_TOTAL_{pay_doc_no}")
                continue
            if current_c_paid_amount is None:
                errors.append(f"MISSING_C_PAID_AMOUNT_{pay_doc_no}")
                continue

            (
                line_factor,
                line_factor_date,
                line_factor_source,
                line_base_currency,
                line_invoice_currency,
            ) = _resolve_line_currency_factor(
                source,
                rates,
                default_base_currency,
            )

            if line_factor in (None, 0):
                errors.append(
                    f"NO_LINE_CURENCY_FACTOR_{pay_doc_no}:"
                    f"BASE={line_base_currency},"
                    f"CONV={line_invoice_currency}"
                )
                continue

            if payment_lookup_factor in (None, 0):
                errors.append(
                    f"NO_PAYMENT_ACCOUNT_FACTOR_{pay_doc_no}:"
                    f"BASE={line_base_currency},"
                    f"ACCOUNT_CURRENCY={account_currency}"
                )
                continue

            # C_PAID_AMOUNT is the current receipt delta in invoice converted
            # currency. ARAP.C_PAID_TOTAL is the cumulative value after adding
            # this delta; it is not a per-receipt input.
            line_c_netamount = current_c_paid_amount
            line_netamount = _calculated_amount(
                line_c_netamount / line_factor
            )
            line_account_amount = _calculated_amount(
                line_netamount * payment_lookup_factor
            )
            conversion_mode = (
                "PAYMENT_EQUALS_INVOICE_CURRENCY"
                if account_currency == line_invoice_currency
                else "PAYMENT_DIFFERS_FROM_INVOICE_CURRENCY"
            )

            line_c_amount = inv_c_doc_total
            line_amount = _calculated_amount(
                line_c_amount / line_factor
            )
            line_c_discount = _calculated_amount(
                line_c_amount - line_c_netamount
            )
            line_discount = _calculated_amount(
                line_amount - line_netamount
            )

            existing_paid_total = _number(
                source.get("PAID_TOTAL_ARAP")
                if "PAID_TOTAL_ARAP" in source.index
                else source.get("PAID_TOTAL"),
                0,
            ) or 0.0
            existing_c_paid_total = _number(
                source.get("C_PAID_TOTAL_ARAP")
                if "C_PAID_TOTAL_ARAP" in source.index
                else source.get("C_PAID_TOTAL"),
                0,
            ) or 0.0

            new_paid_total = _calculated_amount(
                existing_paid_total + line_netamount
            )
            new_c_paid_total = _calculated_amount(
                existing_c_paid_total + line_c_netamount
            )

            calculated.append(
                {
                    "source": source,
                    "LINE_ACCOUNT_AMOUNT": line_account_amount,
                    "ACCOUNT_CURRENCY": account_currency,
                    "LINE_BASE_CURRENCY": line_base_currency,
                    "INVOICE_CURRENCY": line_invoice_currency,
                    "LINE_FACTOR": line_factor,
                    "LINE_FACTOR_DATE": line_factor_date,
                    "LINE_FACTOR_SOURCE": line_factor_source,
                    "CONVERSION_MODE": conversion_mode,
                    "INV_DOC_TOTAL": inv_doc_total,
                    "INV_C_DOC_TOTAL": inv_c_doc_total,
                    "INV_DOC_TOTAL_DIFF": _diff(
                        line_amount,
                        inv_doc_total,
                    ),
                    "EXISTING_ARAP_PAID_TOTAL": existing_paid_total,
                    "EXISTING_ARAP_C_PAID_TOTAL": existing_c_paid_total,
                    "NEW_ARAP_PAID_TOTAL": new_paid_total,
                    "NEW_ARAP_C_PAID_TOTAL": new_c_paid_total,
                    "LINE_AMOUNT": line_amount,
                    "LINE_NETAMOUNT": line_netamount,
                    "LINE_DISCOUNT": line_discount,
                    "LINE_C_AMOUNT": line_c_amount,
                    "LINE_C_NETAMOUNT": line_c_netamount,
                    "LINE_C_DISCOUNT": line_c_discount,
                }
            )

        line_base_values = sorted(
            {
                item["LINE_BASE_CURRENCY"]
                for item in calculated
                if item["LINE_BASE_CURRENCY"]
            }
        )
        base_currency = (
            line_base_values[0]
            if line_base_values
            else default_base_currency
        )
        if len(line_base_values) > 1:
            errors.append("MIXED_BASE_CURRENCY")

        line_conv_values = sorted(
            {
                item["INVOICE_CURRENCY"]
                for item in calculated
                if item["INVOICE_CURRENCY"]
            }
        )
        header_conv = (
            line_conv_values[0]
            if line_conv_values
            else base_currency
        )
        if len(line_conv_values) > 1:
            errors.append("MIXED_CURENCY_CONV_REVIEW")

        applied = _calculated_amount(
            sum(item["LINE_AMOUNT"] or 0 for item in calculated)
        ) or 0.0
        amount = _calculated_amount(
            sum(item["LINE_NETAMOUNT"] or 0 for item in calculated)
        ) or 0.0
        discount = _calculated_amount(
            sum(item["LINE_DISCOUNT"] or 0 for item in calculated)
        ) or 0.0
        c_applied = _calculated_amount(
            sum(item["LINE_C_AMOUNT"] or 0 for item in calculated)
        ) or 0.0
        c_amount = _calculated_amount(
            sum(item["LINE_C_NETAMOUNT"] or 0 for item in calculated)
        ) or 0.0
        c_discount = _calculated_amount(
            sum(item["LINE_C_DISCOUNT"] or 0 for item in calculated)
        ) or 0.0

        conversion_modes = sorted(
            {item["CONVERSION_MODE"] for item in calculated}
        )
        factor_sources = sorted(
            {item["LINE_FACTOR_SOURCE"] for item in calculated}
        )
        factor_values = sorted(
            {
                round(float(item["LINE_FACTOR"]), 9)
                for item in calculated
                if item["LINE_FACTOR"] not in (None, 0)
            }
        )

        account_amount = _calculated_amount(
            sum(item["LINE_ACCOUNT_AMOUNT"] or 0 for item in calculated)
        ) or 0.0

        zero_effective_amount = abs(amount) <= AMOUNT_TOLERANCE
        effective_account_factor = (
            None
            if zero_effective_amount
            else _calculated_amount(account_amount / amount)
        )
        effective_currency_factor = (
            None
            if zero_effective_amount
            else _calculated_amount(c_amount / amount)
        )

        # ACCOUNT_FACTOR is sourced from the base-to-payment TBLCONV lookup.
        # ACCOUNT_AMOUNT / AMOUNT is retained as a validation relationship.
        account_factor = _calculated_amount(payment_lookup_factor)
        account_factor_date = payment_lookup_factor_date
        account_rate_status = (
            "TBLCONV_PAYMENT_LOOKUP:" + payment_lookup_status
        )
        if account_factor in (None, 0):
            errors.append("NO_ACCOUNT_FACTOR")

        if len(factor_values) == 1:
            currency_factor = factor_values[0]
            common_factor_dates = [
                item["LINE_FACTOR_DATE"]
                for item in calculated
                if item["LINE_FACTOR_DATE"] is not None
            ]
            currency_date = (
                common_factor_dates[0]
                if common_factor_dates
                else check_date
            )
            currency_rate_status = "COMMON_LINE_FACTOR"
        elif not zero_effective_amount:
            currency_factor = effective_currency_factor
            currency_date = check_date
            currency_rate_status = "EFFECTIVE_FROM_LINES"
        else:
            currency_factor = None
            currency_date = None
            currency_rate_status = "UNAVAILABLE_ZERO_MIXED_FACTOR_AMOUNT"
            errors.append(
                "OUT_OF_SCOPE_ZERO_PAYMENT_RECEIPT:"
                f"ACCOUNT_AMOUNT={account_amount},"
                f"SUM_LINE_NETAMOUNT={amount},"
                f"SUM_LINE_C_NETAMOUNT={c_amount}"
            )

        if currency_factor in (None, 0):
            errors.append("NO_CURENCY_FACTOR")

        generated_lines: list[dict[str, Any]] = []
        for item in calculated:
            source = item["source"]
            line_row = _init_row(LINE_SCHEMA, LINE_DEFAULTS)
            line_row.update(
                {
                    "DOC_NO": doc_no,
                    "CHECK_NO": check_no,
                    "PAY_USER_DOC": _user_doc(source.get("USER_DOC")),
                    "PAY_DOC_CATEGORY": "IV",
                    "LINE": allocator.next(
                        "R2_LINE_NO",
                        companyno,
                        counter_division,
                    ),
                    "PAY_DOC_NO": int(source.get("DOC_NO")),
                    "PAY_APPLY": check_date,
                    "PAY_DOC_DATE": (
                        (_date(source.get("DOC_DATE")) or check_date)
                        .normalize()
                        + timedelta(days=2)
                    ),
                    "PAY_TYPE": _text(source.get("DOC_TYPE")) or "I",
                    "DISCOUNT": item["LINE_DISCOUNT"],
                    "NETAMOUNT": item["LINE_NETAMOUNT"],
                    "AMOUNT": item["LINE_AMOUNT"],
                    "ACCTNO": (
                        _text(source.get("ACCTNO_INPUT"))
                        or _text(source.get("ACCTNO_ARAP"))
                    ),
                    "SUBC": source.get("SUBC"),
                    "GL_CODE_AMT": "110101-00",
                    "GL_CODE_DISC": (
                        _text(source.get("GL_CODE_DISC"))
                        or "750201-00"
                    ),
                    "COMPANYNO": source.get("COMPANYNO"),
                    "DIVISION": source.get("DIVISION"),
                    "DEPART": source.get("DEPART"),
                    "C_DISCOUNT": item["LINE_C_DISCOUNT"],
                    "C_NETAMOUNT": item["LINE_C_NETAMOUNT"],
                    "C_AMOUNT": item["LINE_C_AMOUNT"],
                    "ADDED_USR": user,
                    "ADDED_DTE": process_ts,
                }
            )
            generated_lines.append(line_row)
            line_rows.append(line_row)

            arap_row: dict[str, Any] = {}
            for column in ARAP_SCHEMA:
                arap_column = f"{column}_ARAP"
                arap_row[column] = (
                    source.get(arap_column)
                    if arap_column in source.index
                    else source.get(column)
                )
            arap_row.update(
                {
                    "USER_DOC": _user_doc(source.get("USER_DOC")),
                    "DOC_TOTAL": item["INV_DOC_TOTAL"],
                    "C_DOC_TOTAL": item["INV_C_DOC_TOTAL"],
                    "PAID_TOTAL": item["NEW_ARAP_PAID_TOTAL"],
                    "C_PAID_TOTAL": item["NEW_ARAP_C_PAID_TOTAL"],
                    "UPDATED_USR": user,
                    "UPDATED_DTE": process_ts,
                }
            )
            arap_rows.append(arap_row)

        matched_divisions = [
            value
            for value in matched.get(
                "DIVISION",
                pd.Series(dtype=object),
            ).tolist()
            if not pd.isna(value)
        ]
        matched_departs = [
            value
            for value in matched.get(
                "DEPART",
                pd.Series(dtype=object),
            ).tolist()
            if not pd.isna(value)
        ]

        hdr_company = bank.get("COMPANY")
        if hdr_company is None or _text(hdr_company) == "":
            hdr_company = companyno

        hdr_division = bank.get("DIVISION")
        if hdr_division is None or _text(hdr_division) == "":
            hdr_division = (
                matched_divisions[0]
                if matched_divisions
                else bank.get("DIVISION")
            )

        hdr_depart = bank.get("DEPART")
        if hdr_depart is None or _text(hdr_depart) == "":
            hdr_depart = (
                matched_departs[0]
                if matched_departs
                else bank.get("DEPART")
            )

        hdr_row = _init_row(HDR_SCHEMA, HDR_DEFAULTS)
        hdr_row.update(
            {
                "DOC_NO": doc_no,
                "CHECK_NO": check_no,
                "CHECK_COUNTER": check_counter,
                "BATCH_NO": batch_no,
                "BANK_ID": bank_id,
                "ACCOUNT_NO": bank.get("ACCOUNT_NO"),
                "ACCTNO": acctno,
                "SUBC": header_subc,
                "CUST_VEND": "C",
                "CHECK_DATE": check_date,
                "PAYMENT_DATE": check_date,
                "NOTE": note_value,
                "PAYEE": customer.get("NAME"),
                "GL_CODE": bank.get("GL_CODE"),
                "COMPANY": hdr_company,
                "DIVISION": hdr_division,
                "DEPART": hdr_depart,
                "CURENCY_STATE": hdr_currency_state,
                "CURENCY_BASE": base_currency,
                "CURENCY_CONV": header_conv,
                "CURENCY_FACTOR": currency_factor,
                "CURENCY_DATE": currency_date,
                "DISCOUNT": discount,
                "NETAMOUNT": amount,
                "AMOUNT": amount,
                "APPLIED": applied,
                "C_AMOUNT": c_amount,
                "C_DISCOUNT": c_discount,
                "C_NETAMOUNT": c_amount,
                "C_APPLIED": c_applied,
                "ACCOUNT_CURRENCY": account_currency,
                "ACCOUNT_AMOUNT": account_amount,
                "ACCOUNT_FACTOR": account_factor,
                "ACCOUNT_FACTOR_DATE": account_factor_date,
                "ADDED_USR": user,
                "ADDED_DTE": process_ts,
            }
        )
        hdr_rows.append(hdr_row)

        sum_line_amount = sum(
            _number(row.get("AMOUNT"), 0)
            for row in generated_lines
        )
        sum_line_net = sum(
            _number(row.get("NETAMOUNT"), 0)
            for row in generated_lines
        )
        sum_line_discount = sum(
            _number(row.get("DISCOUNT"), 0)
            for row in generated_lines
        )
        sum_line_c_amount = sum(
            _number(row.get("C_AMOUNT"), 0)
            for row in generated_lines
        )
        sum_line_c_net = sum(
            _number(row.get("C_NETAMOUNT"), 0)
            for row in generated_lines
        )
        sum_line_c_discount = sum(
            _number(row.get("C_DISCOUNT"), 0)
            for row in generated_lines
        )

        line_payment_detail = " | ".join(
            (
                f"PAY_DOC_NO_DERIVED={int(item['source']['DOC_NO'])}:"
                f"PAY_USER_DOC_INPUT={_user_doc(item['source'].get('PAY_USER_DOC'))},"
                f"PAY_USER_DOC_ARAP={_user_doc(item['source'].get('USER_DOC'))},"
                f"C_PAID_AMOUNT_INPUT={item['LINE_C_NETAMOUNT']},"
                f"LINE_ACCOUNT_AMOUNT_CALCULATED={item['LINE_ACCOUNT_AMOUNT']},"
                f"ACCOUNT_CURRENCY={item['ACCOUNT_CURRENCY']},"
                f"BASE={item['LINE_BASE_CURRENCY']},"
                f"INVOICE_CURRENCY={item['INVOICE_CURRENCY']},"
                f"LINE_FACTOR={item['LINE_FACTOR']},"
                f"FACTOR_SOURCE={item['LINE_FACTOR_SOURCE']},"
                f"MODE={item['CONVERSION_MODE']},"
                f"LINE_C_NETAMOUNT={item['LINE_C_NETAMOUNT']},"
                f"LINE_NETAMOUNT={item['LINE_NETAMOUNT']},"
                f"ARAP_C_PAID_TOTAL:"
                f"{item['EXISTING_ARAP_C_PAID_TOTAL']}"
                f"->{item['NEW_ARAP_C_PAID_TOTAL']},"
                f"ARAP_PAID_TOTAL:"
                f"{item['EXISTING_ARAP_PAID_TOTAL']}"
                f"->{item['NEW_ARAP_PAID_TOTAL']}"
            )
            for item in calculated
        )

        expected_account_factor = effective_account_factor
        expected_currency_factor = effective_currency_factor

        amount_chain_checks = {
            "HDR_ACCOUNT_AMOUNT_VS_SUM_LINE_ACCOUNT_AMOUNT": abs(
                _number(hdr_row.get("ACCOUNT_AMOUNT"), 0)
                - sum(
                    _number(item.get("LINE_ACCOUNT_AMOUNT"), 0)
                    for item in calculated
                )
            ) <= AMOUNT_TOLERANCE,
            "HDR_APPLIED_VS_SUM_LINE_AMOUNT": abs(
                _number(hdr_row.get("APPLIED"), 0)
                - sum_line_amount
            ) <= AMOUNT_TOLERANCE,
            "HDR_AMOUNT_VS_SUM_LINE_NETAMOUNT": abs(
                _number(hdr_row.get("AMOUNT"), 0)
                - sum_line_net
            ) <= AMOUNT_TOLERANCE,
            "HDR_DISCOUNT_VS_SUM_LINE_DISCOUNT": abs(
                _number(hdr_row.get("DISCOUNT"), 0)
                - sum_line_discount
            ) <= AMOUNT_TOLERANCE,
            "HDR_C_APPLIED_VS_SUM_LINE_C_AMOUNT": abs(
                _number(hdr_row.get("C_APPLIED"), 0)
                - sum_line_c_amount
            ) <= AMOUNT_TOLERANCE,
            "HDR_C_AMOUNT_VS_SUM_LINE_C_NETAMOUNT": abs(
                _number(hdr_row.get("C_AMOUNT"), 0)
                - sum_line_c_net
            ) <= AMOUNT_TOLERANCE,
            "HDR_C_DISCOUNT_VS_SUM_LINE_C_DISCOUNT": abs(
                _number(hdr_row.get("C_DISCOUNT"), 0)
                - sum_line_c_discount
            ) <= AMOUNT_TOLERANCE,
            "HDR_NETAMOUNT_VS_AMOUNT": abs(
                _number(hdr_row.get("NETAMOUNT"), 0)
                - _number(hdr_row.get("AMOUNT"), 0)
            ) <= AMOUNT_TOLERANCE,
            "HDR_C_NETAMOUNT_VS_C_AMOUNT": abs(
                _number(hdr_row.get("C_NETAMOUNT"), 0)
                - _number(hdr_row.get("C_AMOUNT"), 0)
            ) <= AMOUNT_TOLERANCE,
            "HDR_ACCOUNT_FACTOR_EFFECTIVE": (
                True
                if expected_account_factor is None
                else abs(
                    _number(hdr_row.get("ACCOUNT_FACTOR"), 0)
                    - expected_account_factor
                ) <= 1e-6
            ),
            "HDR_CURENCY_FACTOR_EFFECTIVE": (
                True
                if expected_currency_factor is None
                else abs(
                    _number(hdr_row.get("CURENCY_FACTOR"), 0)
                    - expected_currency_factor
                ) <= 1e-6
            ),
        }

        failed_amount_checks = [
            name
            for name, passed in amount_chain_checks.items()
            if not passed
        ]
        if failed_amount_checks:
            errors.append(
                "AMOUNT_CHAIN_MISMATCH:"
                + ",".join(failed_amount_checks)
            )

        required_hdr = [
            "DOC_NO",
            "DOC_TYPE",
            "DOC_CATEGORY",
            "CHECK_NO",
            "DOC_STATUS",
            "CHECK_COUNTER",
            "BANK_ID",
            "ACCOUNT_NO",
            "ACCTNO",
            "SUBC",
            "CHECK_DATE",
            "PAYEE",
            "GL_CODE",
            "COMPANY",
            "CURENCY_BASE",
            "CURENCY_CONV",
            "CURENCY_FACTOR",
            "AMOUNT",
            "APPLIED",
            "C_AMOUNT",
            "C_APPLIED",
            "ACCOUNT_CURRENCY",
            "ACCOUNT_AMOUNT",
            "ACCOUNT_FACTOR",
            "ADDED_USR",
            "ADDED_DTE",
        ]
        missing_hdr = [
            field
            for field in required_hdr
            if hdr_row.get(field) is None
            or _text(hdr_row.get(field)) == ""
        ]
        if missing_hdr:
            errors.append(
                "MISSING_HDR_FIELDS:" + ",".join(missing_hdr)
            )

        unique_errors = list(dict.fromkeys(errors))
        reference_messages: list[str] = []
        out_of_scope_messages: list[str] = []
        core_logic_messages: list[str] = []

        for message in unique_errors:
            if message.startswith((
                "CUSTOMER_LOOKUP_NO_MATCH:",
                "CUSTOMER_LOOKUP_NOT_UNIQUE:",
            )):
                reference_messages.append(message)
                continue

            if message.startswith(
                "OUT_OF_SCOPE_ZERO_PAYMENT_RECEIPT:"
            ):
                out_of_scope_messages.append(message)
                continue

            if message.startswith("MISSING_HDR_FIELDS:"):
                missing_fields = {
                    field.strip()
                    for field in message.split(":", 1)[1].split(",")
                    if field.strip()
                }
                reference_fields = missing_fields & {"PAYEE"}
                zero_factor_fields = (
                    missing_fields
                    & {"ACCOUNT_FACTOR", "CURENCY_FACTOR"}
                    if out_of_scope_messages
                    else set()
                )
                remaining_fields = (
                    missing_fields
                    - reference_fields
                    - zero_factor_fields
                )
                if reference_fields:
                    reference_messages.append(
                        "MISSING_HDR_FIELDS:"
                        + ",".join(sorted(reference_fields))
                    )
                if zero_factor_fields:
                    out_of_scope_messages.append(
                        "MISSING_HDR_FIELDS:"
                        + ",".join(sorted(zero_factor_fields))
                    )
                if remaining_fields:
                    core_logic_messages.append(
                        "MISSING_HDR_FIELDS:"
                        + ",".join(sorted(remaining_fields))
                    )
                continue

            core_logic_messages.append(message)

        logic_status = (
            "FAIL"
            if core_logic_messages
            else (
                "OUT_OF_SCOPE"
                if out_of_scope_messages
                else "PASS"
            )
        )
        reference_status = (
            "REVIEW" if reference_messages else "PASS"
        )

        validation_rows.append(
            {
                "INPUT_HDR_GROUP_ID": input_hdr_group_id,
                "DOC_NO": doc_no,
                "CHECK_NO": check_no,
                "CHECK_COUNTER": check_counter,
                "BATCH_NO": batch_no,
                "LINE_COUNT": len(generated_lines),
                "BANK_CURRENCY_COLUMN": bank_currency_column,
                "BANK_MATCH_COUNT": len(bank_matches),
                "CUSTOMER_LOOKUP_KEYS": (
                    f"ACCTNO={acctno};"
                    f"SUBC={_text(header_subc)};"
                    "CUST_VEND=C"
                ),
                "CUSTOMER_MATCH_COUNT": len(customer_matches),
                "CUSTOMER_UNIQUE_NAME_COUNT":
                    len(unique_customer_names),
                "CUSTOMER_LOOKUP_STATUS": customer_lookup_status,
                "ACCOUNT_RATE_STATUS": account_rate_status,
                "PAYMENT_LOOKUP_FACTOR": payment_lookup_factor,
                "CURENCY_RATE_STATUS": currency_rate_status,
                "CURENCY_STATE_SOURCE": currency_state_source,
                "CURENCY_STATE_USED": hdr_currency_state,
                "ACCOUNT_FACTOR_USED": account_factor,
                "CURENCY_FACTOR_USED": currency_factor,
                "DISTINCT_LINE_FACTOR_COUNT": len(factor_values),
                "LINE_FACTORS": ",".join(
                    str(value) for value in factor_values
                ),
                "LINE_FACTOR_SOURCES": ",".join(factor_sources),
                "PAYMENT_CONVERSION_MODE": ",".join(conversion_modes),
                "LINE_PAYMENT_DETAIL": line_payment_detail,
                "INV_HDR_LINE_COUNT": sum(
                    1
                    for item in calculated
                    if item.get("INV_C_DOC_TOTAL") is not None
                ),
                "GENERATED_HDR_AMOUNT": hdr_row.get("AMOUNT"),
                "SUM_LINE_AMOUNT": sum_line_amount,
                "HDR_APPLIED": hdr_row.get("APPLIED"),
                "APPLIED_DIFF": _diff(
                    hdr_row.get("APPLIED"),
                    sum_line_amount,
                ),
                "SUM_LINE_NETAMOUNT": sum_line_net,
                "HDR_AMOUNT": hdr_row.get("AMOUNT"),
                "LINE_NET_TO_HDR_AMOUNT_DIFF": _diff(
                    hdr_row.get("AMOUNT"),
                    sum_line_net,
                ),
                "SUM_LINE_DISCOUNT": sum_line_discount,
                "HDR_DISCOUNT": hdr_row.get("DISCOUNT"),
                "DISCOUNT_DIFF": _diff(
                    hdr_row.get("DISCOUNT"),
                    sum_line_discount,
                ),
                "SUM_LINE_C_AMOUNT": sum_line_c_amount,
                "HDR_C_APPLIED": hdr_row.get("C_APPLIED"),
                "C_APPLIED_DIFF": _diff(
                    hdr_row.get("C_APPLIED"),
                    sum_line_c_amount,
                ),
                "SUM_LINE_C_NETAMOUNT": sum_line_c_net,
                "HDR_C_AMOUNT": hdr_row.get("C_AMOUNT"),
                "LINE_C_NET_TO_HDR_C_AMOUNT_DIFF": _diff(
                    hdr_row.get("C_AMOUNT"),
                    sum_line_c_net,
                ),
                "SUM_LINE_C_DISCOUNT": sum_line_c_discount,
                "HDR_C_DISCOUNT": hdr_row.get("C_DISCOUNT"),
                "C_DISCOUNT_DIFF": _diff(
                    hdr_row.get("C_DISCOUNT"),
                    sum_line_c_discount,
                ),
                "AMOUNT_CHAIN_DETAIL": (
                    f"HDR_AMOUNT={_number(hdr_row.get('AMOUNT'), 0)};"
                    f"SUM_LINE_NETAMOUNT={sum_line_net};"
                    f"HDR_C_AMOUNT={_number(hdr_row.get('C_AMOUNT'), 0)};"
                    f"SUM_LINE_C_NETAMOUNT={sum_line_c_net};"
                    f"ACCOUNT_FACTOR={account_factor};"
                    f"CURENCY_FACTOR={currency_factor}"
                ),
                "LOGIC_STATUS": logic_status,
                "REFERENCE_STATUS": reference_status,
                "CORE_LOGIC_ERRORS":
                    " | ".join(core_logic_messages),
                "REFERENCE_WARNINGS":
                    " | ".join(reference_messages),
                "OUT_OF_SCOPE_REASON":
                    " | ".join(out_of_scope_messages),
                "READY_TO_INSERT":
                    "Y" if not unique_errors else "N",
                "ERRORS": " | ".join(unique_errors),
            }
        )

    _write_plain_sheets(
        output_path,
        {
            "Validation": pd.DataFrame(validation_rows),
            "CHECK_HDR": pd.DataFrame(hdr_rows, columns=HDR_SCHEMA),
            "CHECK_LINE": pd.DataFrame(line_rows, columns=LINE_SCHEMA),
            "ARAP_UPDATE": pd.DataFrame(arap_rows, columns=ARAP_SCHEMA),
            "COUNTER_UPDATE": allocator.output(),
        },
    )



# -----------------------------------------------------------------------------
# 3. Compare generated output against current database rows, column by column
# -----------------------------------------------------------------------------

COMPARISON_AMOUNT_COLUMNS = {
    "ACCOUNT_AMOUNT",
    "AMOUNT",
    "APPLIED",
    "DISCOUNT",
    "NETAMOUNT",
    "OPEN_BALANCE",
    "C_AMOUNT",
    "C_APPLIED",
    "C_DISCOUNT",
    "C_NETAMOUNT",
    "C_OPEN_BALANCE",
    "DOC_TOTAL",
    "C_DOC_TOTAL",
    "PAID_TOTAL",
    "C_PAID_TOTAL",
    "VOID_TOTAL",
    "C_VOID_TOTAL",
    "BALANCE_TOTAL",
    "C_BALANCE_TOTAL",
    "DISCOUNT_AMT",
    "C_DISCOUNT_AMT",
}

COMPARISON_FACTOR_COLUMNS = {
    "ACCOUNT_FACTOR",
    "CURENCY_FACTOR",
}

COMPARISON_CURRENCY_COLUMNS = {
    "ACCOUNT_CURRENCY",
    "CURENCY_BASE",
    "CURENCY_CONV",
}

COMPARISON_DATE_COLUMNS = {
    "CHECK_DATE",
    "PAYMENT_DATE",
    "PRINTED_ON",
    "CURENCY_DATE",
    "ACCOUNT_FACTOR_DATE",
    "ADDED_DTE",
    "UPDATED_DTE",
    "PAY_APPLY",
    "PAY_DOC_DATE",
    "DOC_DATE",
    "DUE_DATE",
    "BATCH_DATE",
    "PRINT_DATE",
    "POST_GL_DATE",
    "CLEARED_ON",
    "WIRE_PROCESS_DATE",
    "SIGN_1_DATE",
    "SIGN_2_DATE",
    "SIGN_3_DATE",
    "SIGN_4_DATE",
    "SIGN_5_DATE",
}


def _state_map_from_workbook(
    generated_path: str | Path,
) -> dict[tuple[str, int, str], str]:
    try:
        state = _read_optional_sheet(
            generated_path,
            ["OUTPUT_STATE"],
        )
    except Exception:
        return {}

    if state.empty:
        return {}

    state = _normalize_columns(state)
    required = {
        "SHEET",
        "OUTPUT_ROW_NUMBER",
        "COLUMN",
        "VALUE_STATE",
    }
    if not required.issubset(state.columns):
        return {}

    result: dict[tuple[str, int, str], str] = {}
    for _, row in state.iterrows():
        result[
            (
                _text(row["SHEET"]).upper(),
                int(_number(row["OUTPUT_ROW_NUMBER"], 0) or 0),
                _text(row["COLUMN"]).upper(),
            )
        ] = _text(row["VALUE_STATE"]).upper()
    return result


def _output_value_state(
    value: Any,
    explicit_state: str | None,
) -> str:
    if explicit_state in {"NULL", "BLANK"}:
        return explicit_state

    state = _cell_state(value)
    if state in {"NULL", "BLANK"}:
        # Old output files do not have enough information to distinguish
        # NULL from an Excel blank cell.
        return "EMPTY_UNKNOWN"
    return "VALUE"


def _db_value_state(value: Any) -> str:
    return _cell_state(value)


def _try_number(value: Any) -> float | None:
    try:
        return _number(value)
    except (TypeError, ValueError):
        return None


def _normalize_comparison_text(
    value: Any,
    column: str,
) -> str:
    text_value = str(value).rstrip()
    if column in COMPARISON_CURRENCY_COLUMNS:
        return _currency(text_value)
    return text_value


def _compare_output_db_value(
    output_value: Any,
    db_value: Any,
    column: str,
    output_state: str,
    amount_tolerance: float,
    numeric_tolerance: float,
    factor_tolerance: float,
    date_tolerance_seconds: float,
) -> tuple[str, str, Any]:
    db_state = _db_value_state(db_value)

    if output_state == "EMPTY_UNKNOWN":
        if db_state in {"NULL", "BLANK"}:
            return (
                "EMPTY_STATE_UNKNOWN",
                "Older output Excel cannot distinguish NULL from BLANK.",
                None,
            )
        return (
            "MISMATCH",
            f"Output is empty but DB state is {db_state}.",
            None,
        )

    if output_state != "VALUE" or db_state != "VALUE":
        if output_state == db_state:
            return "MATCH", "", None
        return (
            "MISMATCH",
            f"Output state={output_state}; DB state={db_state}.",
            None,
        )

    if column in COMPARISON_DATE_COLUMNS:
        try:
            output_date = pd.Timestamp(output_value)
            db_date = pd.Timestamp(db_value)
            diff_seconds = abs(
                (output_date - db_date).total_seconds()
            )
            status = (
                "MATCH"
                if diff_seconds <= date_tolerance_seconds
                else "MISMATCH"
            )
            return (
                status,
                (
                    ""
                    if status == "MATCH"
                    else f"Date difference {diff_seconds} seconds."
                ),
                diff_seconds,
            )
        except Exception:
            pass

    output_number = _try_number(output_value)
    db_number = _try_number(db_value)
    if output_number is not None and db_number is not None:
        if column in COMPARISON_AMOUNT_COLUMNS:
            tolerance = amount_tolerance
        elif column in COMPARISON_FACTOR_COLUMNS:
            tolerance = factor_tolerance
        else:
            tolerance = numeric_tolerance

        difference = output_number - db_number
        status = (
            "MATCH"
            if abs(difference) <= tolerance
            else "MISMATCH"
        )
        return (
            status,
            (
                ""
                if status == "MATCH"
                else f"Numeric difference {difference}."
            ),
            difference,
        )

    output_text = _normalize_comparison_text(
        output_value,
        column,
    )
    db_text = _normalize_comparison_text(
        db_value,
        column,
    )
    status = "MATCH" if output_text == db_text else "MISMATCH"
    return (
        status,
        (
            ""
            if status == "MATCH"
            else "Text values differ after trimming DB padding."
        ),
        None,
    )


def _key_number(value: Any) -> int | float | None:
    number = _number(value)
    if number is None:
        return None
    return int(number) if float(number).is_integer() else number


def _record_key(
    table_name: str,
    row: pd.Series,
) -> tuple[Any, ...]:
    if table_name == "CHECK_HDR":
        return (_key_number(row.get("DOC_NO")),)

    if table_name == "CHECK_LINE":
        return (
            _key_number(row.get("DOC_NO")),
            _key_number(row.get("LINE")),
        )

    if table_name == "ARAP_UPDATE":
        return (
            _key_number(row.get("DOC_NO")),
            _key_number(row.get("LINE")) or 0,
            _text(row.get("DOC_CATEGORY")),
            _text(row.get("ACCTNO")),
        )

    raise ValueError(f"Unsupported comparison table: {table_name}")


def _format_record_key(
    table_name: str,
    key: tuple[Any, ...],
) -> str:
    if table_name == "CHECK_HDR":
        return f"DOC_NO={key[0]}"
    if table_name == "CHECK_LINE":
        return f"DOC_NO={key[0]},LINE={key[1]}"
    return (
        f"DOC_NO={key[0]},LINE={key[1]},"
        f"DOC_CATEGORY={key[2]},ACCTNO={key[3]}"
    )


def _query_generated_table_rows(
    connection: Any,
    table_name: str,
    generated: pd.DataFrame,
) -> pd.DataFrame:
    sql_table = {
        "CHECK_HDR": "dbo.CHECK_HDR",
        "CHECK_LINE": "dbo.CHECK_LINE",
        "ARAP_UPDATE": "dbo.ACCOUNT_AR_AP",
    }[table_name]

    doc_nos = _sql_values(generated.get(
        "DOC_NO",
        pd.Series(dtype=object),
    ))
    if not doc_nos:
        return pd.DataFrame()

    return _normalize_columns(
        _query_in(
            connection,
            sql_table,
            "DOC_NO",
            doc_nos,
        )
    )


def _build_db_index(
    table_name: str,
    db_rows: pd.DataFrame,
) -> dict[tuple[Any, ...], list[pd.Series]]:
    result: dict[tuple[Any, ...], list[pd.Series]] = {}
    for _, row in db_rows.iterrows():
        key = _record_key(table_name, row)
        result.setdefault(key, []).append(row)
    return result


def _compare_standard_table(
    table_name: str,
    generated: pd.DataFrame,
    db_rows: pd.DataFrame,
    state_map: dict[tuple[str, int, str], str],
    amount_tolerance: float,
    numeric_tolerance: float,
    factor_tolerance: float,
    date_tolerance_seconds: float,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    comparisons: list[dict[str, Any]] = []
    row_summaries: list[dict[str, Any]] = []

    generated = _normalize_columns(generated)
    db_rows = _normalize_columns(db_rows)
    db_index = _build_db_index(table_name, db_rows)
    db_columns = set(db_rows.columns)

    for row_index, output_row in generated.reset_index(drop=True).iterrows():
        excel_row_number = row_index + 2
        key = _record_key(table_name, output_row)
        key_text = _format_record_key(table_name, key)
        matches = db_index.get(key, [])

        mismatch_columns: list[str] = []
        uncertain_columns: list[str] = []
        output_only_columns: list[str] = []

        for column in generated.columns:
            output_value = output_row.get(column)
            explicit_state = state_map.get(
                (
                    table_name,
                    excel_row_number,
                    column,
                )
            )
            output_state = _output_value_state(
                output_value,
                explicit_state,
            )

            if column not in db_columns:
                status = "OUTPUT_ONLY_COLUMN"
                detail = "Column does not exist in the database table."
                db_value = None
                db_state = "NOT_APPLICABLE"
                difference = None
                output_only_columns.append(column)
            elif len(matches) == 0:
                status = "ROW_NOT_FOUND"
                detail = "No database row matched the output key."
                db_value = None
                db_state = "NOT_FOUND"
                difference = None
                mismatch_columns.append(column)
            elif len(matches) > 1:
                status = "ROW_NOT_UNIQUE"
                detail = f"{len(matches)} database rows matched the output key."
                db_value = None
                db_state = "MULTIPLE"
                difference = None
                mismatch_columns.append(column)
            else:
                db_value = matches[0].get(column)
                db_state = _db_value_state(db_value)
                status, detail, difference = _compare_output_db_value(
                    output_value=output_value,
                    db_value=db_value,
                    column=column,
                    output_state=output_state,
                    amount_tolerance=amount_tolerance,
                    numeric_tolerance=numeric_tolerance,
                    factor_tolerance=factor_tolerance,
                    date_tolerance_seconds=date_tolerance_seconds,
                )
                if status == "MISMATCH":
                    mismatch_columns.append(column)
                elif status == "EMPTY_STATE_UNKNOWN":
                    uncertain_columns.append(column)

            comparisons.append(
                {
                    "TABLE": table_name,
                    "RECORD_KEY": key_text,
                    "OUTPUT_ROW_NUMBER": excel_row_number,
                    "COLUMN": column,
                    "OUTPUT_STATE": output_state,
                    "OUTPUT_VALUE": output_value,
                    "DB_STATE": db_state,
                    "DB_VALUE": db_value,
                    "DIFFERENCE": difference,
                    "STATUS": status,
                    "DETAIL": detail,
                }
            )

        if mismatch_columns:
            row_status = "MISMATCH"
        elif uncertain_columns:
            row_status = "EMPTY_STATE_UNKNOWN"
        else:
            row_status = "MATCH"

        row_summaries.append(
            {
                "TABLE": table_name,
                "RECORD_KEY": key_text,
                "OUTPUT_ROW_NUMBER": excel_row_number,
                "DB_MATCH_COUNT": len(matches),
                "ROW_STATUS": row_status,
                "MISMATCH_COLUMNS": ", ".join(mismatch_columns),
                "EMPTY_STATE_UNKNOWN_COLUMNS":
                    ", ".join(uncertain_columns),
                "OUTPUT_ONLY_COLUMNS":
                    ", ".join(output_only_columns),
            }
        )

    return comparisons, row_summaries


def _counter_division_state(
    row: pd.Series,
    sheet_name: str,
    excel_row_number: int,
    state_map: dict[tuple[str, int, str], str],
) -> tuple[str, Any]:
    explicit_state = state_map.get(
        (
            sheet_name,
            excel_row_number,
            "DIVISION",
        )
    )
    state = _output_value_state(
        row.get("DIVISION"),
        explicit_state,
    )
    return state, row.get("DIVISION")


def _compare_counter_table(
    connection: Any,
    generated: pd.DataFrame,
    state_map: dict[tuple[str, int, str], str],
    numeric_tolerance: float,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    comparisons: list[dict[str, Any]] = []
    row_summaries: list[dict[str, Any]] = []
    generated = _normalize_columns(generated)

    for row_index, output_row in generated.reset_index(drop=True).iterrows():
        excel_row_number = row_index + 2
        role = _text(output_row.get("ROLE"))
        companyno = output_row.get("COMPANYNO")
        doc_category, doc_type = COUNTER_ROLE_MAP.get(
            role,
            ("", ""),
        )
        division_state, division_value = _counter_division_state(
            output_row,
            "COUNTER_UPDATE",
            excel_row_number,
            state_map,
        )

        sql = """
            SELECT *
            FROM dbo.COUNTERSTBL
            WHERE RTRIM(DOC_CATEGORY) = ?
              AND RTRIM(DOC_TYPE) = ?
              AND COMPANYNO = ?
        """
        params: list[Any] = [
            doc_category,
            doc_type,
            companyno,
        ]

        if division_state == "NULL":
            sql += " AND DIVISION IS NULL"
        elif division_state == "BLANK":
            sql += (
                " AND DIVISION IS NOT NULL"
                " AND RTRIM(DIVISION) = ''"
            )
        else:
            sql += " AND RTRIM(DIVISION) = ?"
            params.append(_text(division_value))

        matches = _normalize_columns(
            _read_sql_query(connection, sql, params)
        )
        key_text = (
            f"ROLE={role},COMPANYNO={companyno},"
            f"DIVISION_STATE={division_state}"
        )

        mismatch_columns: list[str] = []
        output_only_columns: list[str] = []

        counter_values = (
            matches["COUNTER"].dropna().tolist()
            if "COUNTER" in matches.columns
            else []
        )
        unique_counter_values = sorted(
            {
                _key_number(value)
                for value in counter_values
            }
        )

        counter_after = output_row.get("COUNTER_AFTER")
        if len(matches) == 0:
            mapped_status = "ROW_NOT_FOUND"
            mapped_detail = "No COUNTERSTBL row matched."
            db_counter = None
        elif len(unique_counter_values) > 1:
            mapped_status = "ROW_NOT_UNIQUE"
            mapped_detail = (
                "Matched rows contain different COUNTER values."
            )
            db_counter = None
        else:
            db_counter = (
                unique_counter_values[0]
                if unique_counter_values
                else None
            )
            mapped_status, mapped_detail, _ = (
                _compare_output_db_value(
                    output_value=counter_after,
                    db_value=db_counter,
                    column="COUNTER",
                    output_state="VALUE",
                    amount_tolerance=numeric_tolerance,
                    numeric_tolerance=numeric_tolerance,
                    factor_tolerance=numeric_tolerance,
                    date_tolerance_seconds=0,
                )
            )

        mapped_columns = {
            "DOC_CATEGORY": (
                output_row.get("DOC_CATEGORY"),
                doc_category,
            ),
            "DOC_TYPE": (
                output_row.get("DOC_TYPE"),
                doc_type,
            ),
            "COMPANYNO": (
                output_row.get("COMPANYNO"),
                (
                    matches.iloc[0].get("COMPANYNO")
                    if len(matches) >= 1
                    else None
                ),
            ),
            "DIVISION": (
                output_row.get("DIVISION"),
                (
                    matches.iloc[0].get("DIVISION")
                    if len(matches) >= 1
                    else None
                ),
            ),
            "COUNTER_AFTER": (
                counter_after,
                db_counter,
            ),
            "SOURCE_ROW_COUNT": (
                output_row.get("SOURCE_ROW_COUNT"),
                len(matches),
            ),
        }

        for column in generated.columns:
            output_value = output_row.get(column)
            explicit_state = state_map.get(
                (
                    "COUNTER_UPDATE",
                    excel_row_number,
                    column,
                )
            )
            output_state = _output_value_state(
                output_value,
                explicit_state,
            )

            if column not in mapped_columns:
                status = "OUTPUT_ONLY_COLUMN"
                detail = (
                    "Counter simulation metadata has no single DB column."
                )
                db_value = None
                db_state = "NOT_APPLICABLE"
                difference = None
                output_only_columns.append(column)
            elif len(matches) == 0:
                status = "ROW_NOT_FOUND"
                detail = mapped_detail
                db_value = None
                db_state = "NOT_FOUND"
                difference = None
                mismatch_columns.append(column)
            else:
                output_mapped, db_value = mapped_columns[column]
                db_state = _db_value_state(db_value)

                if column == "COUNTER_AFTER":
                    status = mapped_status
                    detail = mapped_detail
                    difference = (
                        None
                        if db_counter is None
                        else _diff(output_mapped, db_counter)
                    )
                elif column == "DIVISION":
                    status, detail, difference = (
                        _compare_output_db_value(
                            output_value=output_mapped,
                            db_value=db_value,
                            column=column,
                            output_state=output_state,
                            amount_tolerance=numeric_tolerance,
                            numeric_tolerance=numeric_tolerance,
                            factor_tolerance=numeric_tolerance,
                            date_tolerance_seconds=0,
                        )
                    )
                else:
                    status, detail, difference = (
                        _compare_output_db_value(
                            output_value=output_mapped,
                            db_value=db_value,
                            column=column,
                            output_state=output_state,
                            amount_tolerance=numeric_tolerance,
                            numeric_tolerance=numeric_tolerance,
                            factor_tolerance=numeric_tolerance,
                            date_tolerance_seconds=0,
                        )
                    )

                if status in {"MISMATCH", "ROW_NOT_UNIQUE"}:
                    mismatch_columns.append(column)

            comparisons.append(
                {
                    "TABLE": "COUNTER_UPDATE",
                    "RECORD_KEY": key_text,
                    "OUTPUT_ROW_NUMBER": excel_row_number,
                    "COLUMN": column,
                    "OUTPUT_STATE": output_state,
                    "OUTPUT_VALUE": output_value,
                    "DB_STATE": db_state,
                    "DB_VALUE": db_value,
                    "DIFFERENCE": difference,
                    "STATUS": status,
                    "DETAIL": detail,
                }
            )

        row_summaries.append(
            {
                "TABLE": "COUNTER_UPDATE",
                "RECORD_KEY": key_text,
                "OUTPUT_ROW_NUMBER": excel_row_number,
                "DB_MATCH_COUNT": len(matches),
                "ROW_STATUS": (
                    "MISMATCH"
                    if mismatch_columns
                    else "MATCH"
                ),
                "MISMATCH_COLUMNS": ", ".join(mismatch_columns),
                "EMPTY_STATE_UNKNOWN_COLUMNS": "",
                "OUTPUT_ONLY_COLUMNS":
                    ", ".join(output_only_columns),
            }
        )

    return comparisons, row_summaries


def _comparison_summaries(
    cell_comparison: pd.DataFrame,
    row_summary: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if cell_comparison.empty:
        return pd.DataFrame(), pd.DataFrame()

    table_summary = (
        cell_comparison.groupby(
            ["TABLE", "STATUS"],
            dropna=False,
        )
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )

    for column in [
        "MATCH",
        "MISMATCH",
        "EMPTY_STATE_UNKNOWN",
        "OUTPUT_ONLY_COLUMN",
        "ROW_NOT_FOUND",
        "ROW_NOT_UNIQUE",
    ]:
        if column not in table_summary.columns:
            table_summary[column] = 0

    if not row_summary.empty:
        row_counts = (
            row_summary.groupby("TABLE")
            .agg(
                OUTPUT_ROW_COUNT=("RECORD_KEY", "count"),
                MATCHED_ROW_COUNT=(
                    "ROW_STATUS",
                    lambda values: sum(value == "MATCH" for value in values),
                ),
                MISMATCH_ROW_COUNT=(
                    "ROW_STATUS",
                    lambda values: sum(
                        value == "MISMATCH"
                        for value in values
                    ),
                ),
                UNCERTAIN_ROW_COUNT=(
                    "ROW_STATUS",
                    lambda values: sum(
                        value == "EMPTY_STATE_UNKNOWN"
                        for value in values
                    ),
                ),
            )
            .reset_index()
        )
        table_summary = table_summary.merge(
            row_counts,
            how="left",
            on="TABLE",
        )

    column_summary = (
        cell_comparison.groupby(
            ["TABLE", "COLUMN", "STATUS"],
            dropna=False,
        )
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )
    return table_summary, column_summary


def compare_generated_output_to_db(
    generated_path: str | Path,
    output_path: str | Path,
    connection_string: str,
    amount_tolerance: float = AMOUNT_TOLERANCE,
    numeric_tolerance: float = 1e-8,
    factor_tolerance: float = FACTOR_TOLERANCE,
    date_tolerance_seconds: float = 120.0,
) -> None:
    try:
        import pyodbc
    except ImportError as exc:
        raise RuntimeError(
            "pyodbc is required. Install it with: pip install pyodbc"
        ) from exc

    generated_sheets: dict[str, pd.DataFrame] = {}
    for sheet_name in [
        "CHECK_HDR",
        "CHECK_LINE",
        "ARAP_UPDATE",
        "COUNTER_UPDATE",
    ]:
        generated_sheets[sheet_name] = _read_optional_sheet(
            generated_path,
            [sheet_name],
        )

    state_map = _state_map_from_workbook(generated_path)
    comparisons: list[dict[str, Any]] = []
    row_summaries: list[dict[str, Any]] = []

    with pyodbc.connect(connection_string) as connection:
        for table_name in [
            "CHECK_HDR",
            "CHECK_LINE",
            "ARAP_UPDATE",
        ]:
            generated = generated_sheets[table_name]
            if generated.empty:
                continue

            db_rows = _query_generated_table_rows(
                connection,
                table_name,
                _normalize_columns(generated),
            )
            table_comparisons, table_rows = (
                _compare_standard_table(
                    table_name=table_name,
                    generated=generated,
                    db_rows=db_rows,
                    state_map=state_map,
                    amount_tolerance=amount_tolerance,
                    numeric_tolerance=numeric_tolerance,
                    factor_tolerance=factor_tolerance,
                    date_tolerance_seconds=date_tolerance_seconds,
                )
            )
            comparisons.extend(table_comparisons)
            row_summaries.extend(table_rows)

        counter_generated = generated_sheets["COUNTER_UPDATE"]
        if not counter_generated.empty:
            counter_comparisons, counter_rows = (
                _compare_counter_table(
                    connection=connection,
                    generated=counter_generated,
                    state_map=state_map,
                    numeric_tolerance=numeric_tolerance,
                )
            )
            comparisons.extend(counter_comparisons)
            row_summaries.extend(counter_rows)

    cell_comparison = pd.DataFrame(comparisons)
    row_summary = pd.DataFrame(row_summaries)
    table_summary, column_summary = _comparison_summaries(
        cell_comparison,
        row_summary,
    )

    _write_plain_sheets(
        output_path,
        {
            "Table_Summary": table_summary,
            "Row_Summary": row_summary,
            "Column_Summary": column_summary,
            "Cell_Comparison": cell_comparison,
        },
    )


# -----------------------------------------------------------------------------
# 4. Validate historical records directly from the database
# -----------------------------------------------------------------------------

def _parse_doc_no_list(value: str | None) -> list[int]:
    if not value:
        return []

    result: list[int] = []
    for token in value.split(","):
        token = token.strip()
        if not token:
            continue
        try:
            result.append(int(token))
        except ValueError as exc:
            raise ValueError(
                f"Invalid DOC_NO in --doc-no: {token}"
            ) from exc
    return result


def _historical_scope_doc_nos(
    connection: Any,
    top: int,
    doc_nos: list[int],
    date_from: str | None,
    date_to: str | None,
    currency_mode: str,
) -> list[int]:
    if doc_nos:
        return sorted(set(doc_nos))

    if top <= 0:
        raise ValueError("--top must be greater than 0.")

    where = [
        "RTRIM(H.DOC_TYPE) = 'AR'",
        "RTRIM(H.DOC_CATEGORY) = 'R2'",
        "EXISTS ("
        " SELECT 1"
        " FROM dbo.CHECK_LINE L"
        " WHERE L.DOC_NO = H.DOC_NO"
        "   AND RTRIM(L.DOC_CATEGORY) = RTRIM(H.DOC_CATEGORY)"
        "   AND RTRIM(L.PAY_DOC_CATEGORY) = 'IV'"
        ")",
    ]
    params: list[Any] = []

    if date_from:
        where.append("H.CHECK_DATE >= ?")
        params.append(date_from)
    if date_to:
        where.append("H.CHECK_DATE < DATEADD(day, 1, CAST(? AS date))")
        params.append(date_to)

    normalized_account = (
        "CASE "
        "WHEN UPPER(RTRIM(H.ACCOUNT_CURRENCY)) IN ('RMB','CNY') "
        "THEN 'CNY' "
        "ELSE UPPER(RTRIM(H.ACCOUNT_CURRENCY)) END"
    )
    normalized_conv = (
        "CASE "
        "WHEN UPPER(RTRIM(H.CURENCY_CONV)) IN ('RMB','CNY') "
        "THEN 'CNY' "
        "ELSE UPPER(RTRIM(H.CURENCY_CONV)) END"
    )

    if currency_mode == "same":
        where.append(f"{normalized_account} = {normalized_conv}")
    elif currency_mode == "cross":
        where.append(f"{normalized_account} <> {normalized_conv}")

    sql = f"""
        SELECT TOP {int(top)}
            H.DOC_NO
        FROM dbo.CHECK_HDR H
        WHERE {" AND ".join(where)}
        ORDER BY H.CHECK_DATE DESC, H.DOC_NO DESC
    """
    scope = _read_sql_query(connection, sql, params)
    if scope.empty:
        return []

    return [
        int(value)
        for value in scope["DOC_NO"].dropna().tolist()
    ]


def validate_existing_chain_from_db(
    output_path: str | Path,
    connection_string: str,
    top: int = 1000,
    doc_no: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    currency_mode: str = "all",
) -> None:
    """
    Validate historical records directly from SQL Server.

    Core CHECK_HDR and CHECK_LINE validation uses the complete line set for
    each selected receipt. ARAP cumulative comparison uses a separate,
    broader payment-history query and is reported as diagnostic only.
    """
    try:
        import pyodbc
    except ImportError as exc:
        raise RuntimeError(
            "pyodbc is required. Install it with: pip install pyodbc"
        ) from exc

    requested_doc_nos = _parse_doc_no_list(doc_no)

    with pyodbc.connect(connection_string) as connection:
        focus_doc_nos = _historical_scope_doc_nos(
            connection=connection,
            top=top,
            doc_nos=requested_doc_nos,
            date_from=date_from,
            date_to=date_to,
            currency_mode=currency_mode,
        )

        if not focus_doc_nos:
            raise ValueError(
                "No supported AR/R2/IV historical receipts matched "
                "the requested DB scope."
            )

        # Retrieve every R2 line for Header aggregation/classification.
        # The confirmed generation scope remains PAY_DOC_CATEGORY='IV',
        # while AD/MS lines are reported as out-of-scope historical cases.
        focus_all_lines = _query_in(
            connection,
            "dbo.CHECK_LINE",
            "DOC_NO",
            focus_doc_nos,
            extra_where="RTRIM(DOC_CATEGORY) = 'R2'",
        )
        focus_all_lines = _normalize_columns(focus_all_lines)
        focus_lines = focus_all_lines[
            focus_all_lines["PAY_DOC_CATEGORY"]
            .map(_text)
            .eq("IV")
        ].copy()

        focus_hdr = _query_in(
            connection,
            "dbo.CHECK_HDR",
            "DOC_NO",
            focus_doc_nos,
            extra_where=(
                "RTRIM(DOC_TYPE) = 'AR' "
                "AND RTRIM(DOC_CATEGORY) = 'R2'"
            ),
        )
        focus_hdr = _normalize_columns(focus_hdr)

        pay_doc_nos = _sql_values(
            focus_lines.get(
                "PAY_DOC_NO",
                pd.Series(dtype=object),
            )
        )
        if not pay_doc_nos:
            raise ValueError(
                "The selected receipts contain no IV CHECK_LINE rows."
            )

        arap = _query_in(
            connection,
            "dbo.ACCOUNT_AR_AP",
            "DOC_NO",
            pay_doc_nos,
            extra_where=(
                "ISNULL(LINE, 0) = 0 "
                "AND RTRIM(DOC_CATEGORY) = 'IV'"
            ),
        )
        arap = _normalize_columns(arap)

        # Broad invoice payment history for ARAP diagnostics. Do not restrict
        # CHECK_LINE.DOC_CATEGORY to R2 here because ARAP totals may reflect
        # other payment/reversal categories.
        history_lines = _query_in(
            connection,
            "dbo.CHECK_LINE",
            "PAY_DOC_NO",
            pay_doc_nos,
            extra_where="RTRIM(PAY_DOC_CATEGORY) = 'IV'",
        )
        history_lines = _normalize_columns(history_lines)

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_dir_path = Path(temp_dir)
        arap_path = temp_dir_path / "arap.xlsx"
        focus_line_path = temp_dir_path / "focus_check_line.xlsx"
        all_line_path = temp_dir_path / "all_receipt_lines.xlsx"
        focus_hdr_path = temp_dir_path / "focus_check_hdr.xlsx"
        history_line_path = temp_dir_path / "arap_history_line.xlsx"

        arap.to_excel(arap_path, index=False)
        focus_lines.to_excel(focus_line_path, index=False)
        focus_all_lines.to_excel(all_line_path, index=False)
        focus_hdr.to_excel(focus_hdr_path, index=False)
        history_lines.to_excel(history_line_path, index=False)

        validate_existing_chain(
            arap_path=arap_path,
            check_line_path=focus_line_path,
            check_hdr_path=focus_hdr_path,
            output_path=output_path,
            arap_history_line_path=history_line_path,
            header_line_path=all_line_path,
        )

    scope_rows = pd.DataFrame(
        {
            "FOCUS_DOC_NO": focus_doc_nos,
            "CURRENCY_MODE": currency_mode,
            "DATE_FROM": date_from,
            "DATE_TO": date_to,
            "CORE_LINE_SCOPE": "COMPLETE_SELECTED_RECEIPT",
            "ARAP_SCOPE": "ALL_IV_PAYMENT_LINES_FOUND_FOR_FOCUS_INVOICES",
        }
    )

    from openpyxl import load_workbook

    workbook = load_workbook(output_path)
    if "DB_Scope" in workbook.sheetnames:
        del workbook["DB_Scope"]
    worksheet = workbook.create_sheet("DB_Scope")

    for column_index, column_name in enumerate(
        scope_rows.columns,
        start=1,
    ):
        worksheet.cell(1, column_index, column_name)

    for row_index, row in enumerate(
        scope_rows.itertuples(index=False),
        start=2,
    ):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row_index, column_index, value)

    worksheet.freeze_panes = "A2"
    _plain_header(worksheet, 1)
    _resize_plain(worksheet)
    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="P2000 AR Fund In Excel validator and generator")
    sub = parser.add_subparsers(dest="command", required=True)

    validate = sub.add_parser("validate", help="Validate existing ARAP/CHECK_LINE/CHECK_HDR Excel exports")
    validate.add_argument("--arap", required=True)
    validate.add_argument("--check-line", required=True)
    validate.add_argument("--check-hdr", required=True)
    validate.add_argument("--output", required=True)

    validate_db = sub.add_parser(
        "validate-db",
        help=(
            "Query historical CHECK_HDR, CHECK_LINE, and ACCOUNT_AR_AP "
            "directly from the database and validate the confirmed "
            "amount chain."
        ),
    )
    validate_db.add_argument("--output", required=True)
    validate_db.add_argument(
        "--env-file",
        default=".env",
        help="Path to the .env file containing P2000 SQL Login settings.",
    )
    validate_db.add_argument(
        "--top",
        type=int,
        default=1000,
        help="Number of recent supported receipts to select.",
    )
    validate_db.add_argument(
        "--doc-no",
        required=False,
        help="Optional comma-separated CHECK_HDR DOC_NO values.",
    )
    validate_db.add_argument("--date-from", required=False)
    validate_db.add_argument("--date-to", required=False)
    validate_db.add_argument(
        "--currency-mode",
        choices=["all", "same", "cross"],
        default="all",
    )

    generate = sub.add_parser(
        "generate",
        help="Generate full rows from the simple one-sheet input",
    )
    generate.add_argument("--input", required=True)
    generate.add_argument("--output", required=True)
    generate.add_argument(
        "--env-file",
        default=".env",
        help="Path to the .env file containing P2000 SQL Login settings.",
    )
    generate.add_argument("--user", default="ILC")
    generate.add_argument("--process-time", required=False)

    compare_db = sub.add_parser(
        "compare-db",
        help=(
            "Compare every generated output column against current "
            "database rows."
        ),
    )
    compare_db.add_argument("--generated", required=True)
    compare_db.add_argument("--output", required=True)
    compare_db.add_argument(
        "--env-file",
        default=".env",
        help="Path to the .env file containing P2000 SQL Login settings.",
    )
    compare_db.add_argument(
        "--amount-tolerance",
        type=float,
        default=AMOUNT_TOLERANCE,
    )
    compare_db.add_argument(
        "--numeric-tolerance",
        type=float,
        default=1e-8,
    )
    compare_db.add_argument(
        "--factor-tolerance",
        type=float,
        default=FACTOR_TOLERANCE,
    )
    compare_db.add_argument(
        "--date-tolerance-seconds",
        type=float,
        default=120.0,
    )

    args = parser.parse_args()
    if args.command == "validate":
        validate_existing_chain(
            args.arap,
            args.check_line,
            args.check_hdr,
            args.output,
        )
    elif args.command == "validate-db":
        connection_string = _connection_string_from_env(
            args.env_file
        )
        validate_existing_chain_from_db(
            output_path=args.output,
            connection_string=connection_string,
            top=args.top,
            doc_no=args.doc_no,
            date_from=args.date_from,
            date_to=args.date_to,
            currency_mode=args.currency_mode,
        )
    elif args.command == "generate":
        connection_string = _connection_string_from_env(args.env_file)
        generate_insert_ready_excel(
            args.input,
            args.output,
            connection_string,
            user=args.user,
            process_time=args.process_time,
        )
    else:
        connection_string = _connection_string_from_env(args.env_file)
        compare_generated_output_to_db(
            generated_path=args.generated,
            output_path=args.output,
            connection_string=connection_string,
            amount_tolerance=args.amount_tolerance,
            numeric_tolerance=args.numeric_tolerance,
            factor_tolerance=args.factor_tolerance,
            date_tolerance_seconds=args.date_tolerance_seconds,
        )


if __name__ == "__main__":
    main()
